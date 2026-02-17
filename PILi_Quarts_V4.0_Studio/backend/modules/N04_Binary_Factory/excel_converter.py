
import re
import logging
import os
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Try to import Image for Logo, handle gracefully if missing
try:
    from openpyxl.drawing.image import Image as XLImage
    HAS_IMAGE_SUPPORT = True
except ImportError:
    HAS_IMAGE_SUPPORT = False

logger = logging.getLogger(__name__)

class TeslaExcelConverter:
    """
    Convierte HTML Profesional de Tesla a Excel Nativo ("Gemelo Digital")
    V10: "The Mirror" - 12-Column Grid, Frozen Headers, Rich Text.
    """
    def __init__(self):
        # Paleta Tesla
        self.azul_primario = "0052A3"
        self.azul_secundario = "1E40AF"
        self.azul_claro = "EFF6FF"
        self.gris_fondo = "F9FAFB"
        self.gris_texto = "374151"
        self.blanco = "FFFFFF"
        self.borde_gris = "E5E7EB"
        
        # Fuentes
        self.font_header = Font(name='Calibri', size=11, bold=True, color=self.blanco)
        self.font_title = Font(name='Calibri', size=20, bold=True, color=self.azul_primario)
        self.font_section = Font(name='Calibri', size=12, bold=True, color=self.azul_primario)
        self.font_label = Font(name='Calibri', size=11, bold=True, color=self.azul_secundario)
        self.font_text = Font(name='Calibri', size=11, color=self.gris_texto)

        # Zoning Constants
        self.ZONE_1_END = 9
        self.ZONE_2_START = 10
        self.ZONE_3_START = 17
        
        # Assets
        self.assets_dir = Path(__file__).parent / "templates" / "assets"
        self.logo_path = self.assets_dir / "logo.png"

    def clean_numeric(self, text):
        if not text: return 0.0
        clean = re.sub(r'[^\d.-]', '', text.replace(',', ''))
        try:
            return float(clean)
        except ValueError:
            return 0.0

    def set_border(self, cell, style='thin'):
        side = Side(border_style=style, color=self.borde_gris)
        cell.border = Border(left=side, right=side, top=side, bottom=side)

    def convert_cotizacion(self, soup, ws):
        # 1. Configuración de Columnas (12-Col Grid A-L)
        # A: Spacer (Narrow), B-L: Content
        ws.column_dimensions['A'].width = 2
        for col in range(2, 13): # B to L
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 11 # Proportional ~10-12
            
        # Freeze Panes at Row 10 (Lock Branding)
        ws.freeze_panes = f'A{self.ZONE_2_START}'
        
        # ZONE 1: BRANDING (Rows 1-9)
        self._render_zone_1_branding(soup, ws)
        
        # ZONE 2: CLIENT INFO (Start Row 10)
        curr_row = self.ZONE_2_START
        curr_row = self._render_zone_2_data(soup, ws, curr_row)
        
        # ZONE 3: BODY (Start Row 17+)
        curr_row = max(curr_row, self.ZONE_3_START)
        self._render_zone_3_body(soup, ws, curr_row)

    def _render_zone_1_branding(self, soup, ws):
        """Renderiza Logo, Empresa y Título en filas 1-9"""
        # Logo (B2)
        if HAS_IMAGE_SUPPORT and self.logo_path.exists():
            try:
                img = XLImage(str(self.logo_path))
                # Resize logic if needed, usually generic logo fits well in 2-3 rows space
                ws.add_image(img, 'B2')
            except Exception as e:
                logger.warning(f"Could not load logo: {e}")
                ws['B2'] = "TESLA" 
        else:
             ws.merge_cells('B2:C3')
             c = ws['B2']
             c.value = "TESLA"
             c.fill = PatternFill(start_color=self.azul_primario, end_color=self.azul_primario, fill_type="solid")
             c.font = Font(name='Arial Black', size=24, bold=True, color=self.blanco)
             c.alignment = Alignment(horizontal='center', vertical='center')

        # Company Info (H2 - aligned right in Grid)
        header = soup.find(class_='header')
        if header:
            detalles = header.find(class_='empresa-detalles')
            if detalles:
                info_row = 2
                for div in detalles.find_all('div'):
                    if info_row > 5: break
                    # Merge H to L for address info
                    ws.merge_cells(f'H{info_row}:L{info_row}')
                    cell = ws[f'H{info_row}']
                    cell.value = div.get_text(strip=True)
                    cell.font = Font(size=9, color=self.gris_texto)
                    cell.alignment = Alignment(horizontal='right')
                    info_row += 1

        # Title (B6-B8)
        titulo_div = soup.find(class_='titulo-documento')
        if titulo_div:
            row_t = 6
            h1 = titulo_div.find('h1')
            sub = titulo_div.find(class_='subtitulo-documento')
            num = titulo_div.find(class_='numero-cotizacion')
            
            # Center across full width (B-L)
            if h1:
                ws.merge_cells(f'B{row_t}:L{row_t}')
                c = ws[f'B{row_t}']
                c.value = h1.get_text(strip=True)
                c.font = self.font_title
                c.alignment = Alignment(horizontal='center')
                row_t += 1
            if sub:
                ws.merge_cells(f'B{row_t}:L{row_t}')
                c = ws[f'B{row_t}']
                c.value = sub.get_text(strip=True)
                c.font = Font(size=12, italic=True, color=self.azul_secundario)
                c.alignment = Alignment(horizontal='center')
                row_t += 1
            if num:
                ws.merge_cells(f'B{row_t}:L{row_t}')
                c = ws[f'B{row_t}']
                c.value = num.get_text(strip=True)
                c.font = Font(size=14, bold=True, color=self.azul_secundario)
                c.alignment = Alignment(horizontal='center')

    def _render_zone_2_data(self, soup, ws, start_row):
        """Renderiza Info Boxes, Portada (Informe) o Título/Grid (Proyecto)"""
        row = start_row
        
        # 1. COTIZACIÓN: .info-section with .info-box
        info_section = soup.find(class_='info-section')
        if info_section:
            boxes = info_section.find_all(class_='info-box')
            row_b1 = row
            row_b2 = row
            if len(boxes) > 0: row_b1 = self._render_info_box_content(boxes[0], ws, row, 2, 6)
            if len(boxes) > 1: row_b2 = self._render_info_box_content(boxes[1], ws, row, 8, 12)
            return max(row_b1, row_b2)
            
        # 2. INFORME TÉCNICO: .portada
        portada = soup.find(class_='portada')
        if portada:
            # Title
            h1 = portada.find('h1')
            h2 = portada.find('h2')
            if h1:
                ws.merge_cells(f'B{row}:L{row}')
                c = ws[f'B{row}']
                c.value = h1.get_text(strip=True)
                c.font = Font(name='Calibri', size=24, bold=True, color=self.azul_primario)
                c.alignment = Alignment(horizontal='center')
                row += 1
            if h2:
                ws.merge_cells(f'B{row}:L{row}')
                c = ws[f'B{row}']
                c.value = h2.get_text(strip=True)
                c.font = Font(name='Calibri', size=16, bold=True, color=self.azul_secundario)
                c.alignment = Alignment(horizontal='center')
                row += 2
                
            # Info
            p_info = portada.find(class_='portada-info')
            if p_info:
                for div in p_info.find_all('div'):
                    text = div.get_text(strip=True)
                    ws.merge_cells(f'B{row}:L{row}')
                    c = ws[f'B{row}']
                    c.value = text
                    c.alignment = Alignment(horizontal='center')
                    row += 1
            return row + 1

        # 3. PROYECTO: .titulo + .info-grid
        titulo = soup.find(class_='titulo')
        if titulo:
             h1 = titulo.find('h1')
             sub = titulo.find(class_='subtitulo')
             # Render Title
             if h1:
                ws.merge_cells(f'B{row}:L{row}')
                c = ws[f'B{row}']
                c.value = h1.get_text(strip=True)
                c.font = Font(name='Calibri', size=24, bold=True, color=self.azul_primario)
                c.alignment = Alignment(horizontal='center')
                row += 1
             if sub:
                ws.merge_cells(f'B{row}:L{row}')
                c = ws[f'B{row}']
                c.value = sub.get_text(strip=True)
                c.font = Font(size=14, italic=True, color=self.azul_secundario)
                c.alignment = Alignment(horizontal='center')
                row += 2

        info_grid = soup.find(class_='info-grid')
        if info_grid:
            # Grid of 4 cards
            cards = info_grid.find_all(class_='info-card')
            # Layout: 4 cards in one row? Or 2x2.
            # Excel Columns B-L (11 cols). 4 cards -> 2.75 cols? 
            # Let's do 2 rows of 2.
            
            # Row 1: Card 1 (B-F), Card 2 (H-L)
            current_grid_row = row
            if len(cards) >= 1: self._render_simple_card(cards[0], ws, current_grid_row, 2, 6)
            if len(cards) >= 2: self._render_simple_card(cards[1], ws, current_grid_row, 8, 12)
            
            current_grid_row += 3
            # Row 2: Card 3 (B-F), Card 4 (H-L)
            if len(cards) >= 3: self._render_simple_card(cards[2], ws, current_grid_row, 2, 6)
            if len(cards) >= 4: self._render_simple_card(cards[3], ws, current_grid_row, 8, 12)
            
            return current_grid_row + 3
            
        return row

    def _render_simple_card(self, card, ws, row, c_start, c_end):
        lbl = card.find(class_='info-label')
        val = card.find(class_='info-value')
        
        ws.merge_cells(start_row=row, start_column=c_start, end_row=row, end_column=c_end)
        c = ws.cell(row=row, column=c_start, value=lbl.get_text(strip=True) if lbl else "")
        c.font = self.font_label
        c.alignment = Alignment(horizontal='center')
        
        ws.merge_cells(start_row=row+1, start_column=c_start, end_row=row+1, end_column=c_end)
        c = ws.cell(row=row+1, column=c_start, value=val.get_text(strip=True) if val else "")
        c.font = Font(size=12, bold=True, color=self.azul_primario)
        c.alignment = Alignment(horizontal='center')
        c.border = Border(bottom=Side(style='thick', color=self.azul_primario))

    def _render_zone_3_body(self, soup, ws, start_row):
        """Renderiza el resto del contenido con Rich Text (Universal)"""
        curr_row = start_row
        
        # 0. Resumen Ejecutivo (Reports)
        resumen = soup.find(class_='resumen-ejecutivo')
        if resumen:
            curr_row = self._render_generic_content(resumen, ws, curr_row)
            curr_row += 1

        # 1. Sections Universal (.seccion OR .seccion-completa)
        sections = soup.find_all(['div'], class_=['seccion', 'seccion-completa'])
        
        for section in sections:
            curr_row += 1
            # Section Title
            h2 = section.find('h2')
            if h2:
                ws.merge_cells(f'B{curr_row}:L{curr_row}')
                c = ws[f'B{curr_row}']
                c.value = h2.get_text(strip=True).upper()
                c.font = self.font_section
                c.border = Border(bottom=Side(style='thick', color=self.azul_primario))
                curr_row += 1
            
            # Sub-components check
            crono_box = section.find(class_='cronograma-box')
            garantia_grid = section.find(class_='garantias-grid')
            kpis_grid = section.find(class_='kpis-grid')
            stakeholders = section.find_all(class_='stakeholder-card')
            raci_grid = section.find(class_='raci-grid')
             # Special Tables (Riesgos) - standard table inside section
            tables = section.find_all('table')
            
            if crono_box:
                curr_row = self._render_cronograma(crono_box, ws, curr_row)
            elif garantia_grid:
                curr_row = self._render_garantias(garantia_grid, ws, curr_row)
            elif kpis_grid:
                curr_row = self._render_kpis(kpis_grid, ws, curr_row)
            elif stakeholders:
                for sh in stakeholders:
                    curr_row = self._render_stakeholder(sh, ws, curr_row)
            elif raci_grid:
                 curr_row = self._render_raci_table(raci_grid.find('table'), ws, curr_row)
            elif tables:
                 for tbl in tables:
                     curr_row = self._render_generic_table(tbl, ws, curr_row)
            else:
                # Generic Text Content
                # If section has 'h3', 'p', 'ul', 'div.caja-destacada'
                curr_row = self._render_generic_content_advanced(section, ws, curr_row)

        # 2. Tables Reference (Cotizacion Specific) - kept for retro-compatibility
        tabla_section = soup.find(class_='tabla-section')
        if tabla_section:
            curr_row = self._render_items_table(tabla_section, ws, curr_row)
            totales_section = soup.find(class_='totales-section')
            if totales_section:
                curr_row = self._render_totales(totales_section, ws, curr_row)

        # 3. Footer
        footer = soup.find(class_='footer')
        if footer:
            curr_row += 2
            ws.merge_cells(f'B{curr_row}:L{curr_row}')
            c = ws[f'B{curr_row}']
            c.value = footer.get_text(strip=True, separator=" | ")
            c.font = Font(size=9, color=self.gris_texto)
            c.alignment = Alignment(horizontal='center', wrap_text=True)
            c.border = Border(top=Side(style='medium', color=self.azul_primario))
    
    def _render_generic_content_advanced(self, container, ws, start_row):
        row = start_row
        # Iterate direct children to preserve order
        for elem in container.children:
            if elem.name == 'h2': continue # Already handled
            
            text = elem.get_text(strip=True) if hasattr(elem, 'get_text') else str(elem).strip()
            if not text: continue
            
            if elem.name == 'h3':
                ws.merge_cells(f'B{row}:L{row}')
                c = ws[f'B{row}']
                c.value = text
                c.font = Font(size=12, bold=True, color=self.azul_secundario)
                row += 1
                
            elif elem.name == 'p' or elem.name is None:
                ws.merge_cells(f'B{row}:L{row}')
                c = ws[f'B{row}']
                c.value = text
                c.alignment = Alignment(wrap_text=True, vertical='center')
                row += 1
                
            elif elem.name == 'ul' or elem.name == 'ol':
                for li in elem.find_all('li'):
                    ws.merge_cells(f'B{row}:L{row}')
                    c = ws[f'B{row}']
                    marker = "•" if elem.name == 'ul' else f"{li.parent.index(li)+1}." 
                    # Note: index might be wrong for ol logic, but simple enough for now
                    c.value = f"  {marker} {li.get_text(strip=True)}"
                    c.alignment = Alignment(wrap_text=True, indent=1, vertical='center')
                    row += 1
            
            elif elem.name == 'div' and 'caja-destacada' in elem.get('class', []):
                # Box logic
                h4 = elem.find('h4')
                if h4:
                    ws.merge_cells(f'B{row}:L{row}')
                    c = ws[f'B{row}']
                    c.value = h4.get_text(strip=True)
                    c.font = Font(bold=True, color=self.azul_primario)
                    c.fill = PatternFill(start_color=self.gris_fondo, end_color=self.gris_fondo, fill_type='solid')
                    row += 1
                
                # Render content inside box
                row = self._render_generic_content_advanced(elem, ws, row)
                
        return row

    def _render_kpis(self, grid, ws, start_row):
        row = start_row
        col = 2
        for card in grid.find_all(class_='kpi-card'):
            lbl = card.find(class_='kpi-label').get_text(strip=True)
            val = card.find(class_='kpi-value').get_text(strip=True)
            
            # 2 cols per KPI (B-C, D-E, F-G, H-I, J-K) -> 5 cols.
            # L left empty or spanned.
            if col > 11: break
            
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
            c1 = ws.cell(row=row, column=col, value=lbl)
            c1.alignment = Alignment(horizontal='center')
            c1.font = Font(size=9, color=self.gris_texto)
            
            ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
            c2 = ws.cell(row=row+1, column=col, value=val)
            c2.alignment = Alignment(horizontal='center')
            c2.font = Font(name='Calibri', size=14, bold=True, color=self.azul_primario)
            c2.border = Border(bottom=Side(style='thin', color=self.azul_secundario))
            
            col += 2
        return row + 3

    def _render_stakeholder(self, card, ws, row):
        name = card.find(class_='stakeholder-nombre').get_text(strip=True)
        rol = card.find(class_='stakeholder-rol').get_text(strip=True)
        
        ws.merge_cells(f'B{row}:D{row}')
        c = ws[f'B{row}']
        c.value = name
        c.font = Font(bold=True, color=self.azul_primario)
        
        ws.merge_cells(f'E{row}:L{row}')
        c2 = ws[f'E{row}']
        c2.value = rol
        c2.alignment = Alignment(horizontal='right')
        
        return row + 1

    def _render_generic_table(self, table, ws, start_row):
        # Header
        row = start_row
        thead = table.find('thead')
        if thead:
            headers = [th.get_text(strip=True) for th in thead.find_all('th')]
            col_span = 10 // len(headers) # Distribute 10 cols (B-K)
            
            curr_col = 2
            for h in headers:
                ws.merge_cells(start_row=row, start_column=curr_col, end_row=row, end_column=curr_col+col_span-1)
                c = ws.cell(row=row, column=curr_col, value=h)
                c.fill = PatternFill(start_color=self.azul_primario, fill_type='solid')
                c.font = self.font_header
                curr_col += col_span
            row += 1
            
        tbody = table.find('tbody')
        if tbody:
            for tr in tbody.find_all('tr'):
                cols = tr.find_all('td')
                col_span = 10 // len(cols) if cols else 1
                curr_col = 2
                for td in cols:
                    ws.merge_cells(start_row=row, start_column=curr_col, end_row=row, end_column=curr_col+col_span-1)
                    c = ws.cell(row=row, column=curr_col, value=td.get_text(strip=True))
                    c.alignment = Alignment(wrap_text=True, vertical='center')
                    c.border = Border(bottom=Side(style='thin', color=self.borde_gris))
                    curr_col += col_span
                row += 1
                
        return row + 1
        
    def _render_raci_table(self, table, ws, start_row):
        return self._render_generic_table(table, ws, start_row)
            
    def _render_info_box_content(self, box, ws, start_row, start_col, end_col):
        row = start_row
        h3 = box.find('h3')
        if h3:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
            c = ws.cell(row=row, column=start_col, value=h3.get_text(strip=True))
            c.font = Font(bold=True, color=self.azul_primario)
            c.border = Border(bottom=Side(style='thick', color=self.azul_primario))
            row += 1
            
        for p in box.find_all('p'):
            label_span = p.find(class_='info-label')
            label = label_span.get_text(strip=True) if label_span else ""
            value = p.get_text(strip=True).replace(label, "").strip()
            
            # Label
            ws.cell(row=row, column=start_col, value=label).font = self.font_label
            # Value (Merge remaining cols)
            ws.merge_cells(start_row=row, start_column=start_col+1, end_row=row, end_column=end_col)
            c_val = ws.cell(row=row, column=start_col+1, value=value)
            c_val.alignment = Alignment(horizontal='left', wrap_text=True)
            row += 1
        return row

    def _render_generic_content(self, container, ws, start_row):
        row = start_row
        for elem in container.children:
            text = elem.get_text(strip=True) if hasattr(elem, 'get_text') else str(elem).strip()
            if not text: continue
            
            # Rich Text logic: Wrap everything
            if elem.name == 'p' or elem.name is None:
                ws.merge_cells(f'B{row}:L{row}')
                c = ws[f'B{row}']
                c.value = text
                c.alignment = Alignment(wrap_text=True, vertical='center')
                # Check for "Justificado"? Excel justification is tricky, usually 'left' is cleaner.
                row += 1
            elif elem.name == 'ul':
                for li in elem.find_all('li'):
                    ws.merge_cells(f'B{row}:L{row}')
                    c = ws[f'B{row}']
                    c.value = f"• {li.get_text(strip=True)}"
                    c.alignment = Alignment(wrap_text=True, indent=1, vertical='center')
                    row += 1
        return row

    def _render_items_table(self, section, ws, start_row):
        curr_row = start_row + 2
        ws.merge_cells(f'B{curr_row}:L{curr_row}')
        ws[f'B{curr_row}'] = "DETALLE DE LA COTIZACIÓN"
        ws[f'B{curr_row}'].font = self.font_section
        curr_row += 1
        
        # New Grid Layout:
        # Item: B (1)
        # Desc: C-F (4)
        # Cant: G (1)
        # Und: H (1)
        # PU: I-J (2)
        # Total: K-L (2)
        
        headers = [
            ("ITEM", 2, 2), # Col B
            ("DESCRIPCIÓN", 3, 6), # C-F
            ("CANT.", 7, 7), # G
            ("UNIDAD", 8, 8), # H
            ("P. UNIT.", 9, 10), # I-J
            ("TOTAL", 11, 12) # K-L
        ]
        
        for h_text, start_c, end_c in headers:
            if start_c == end_c:
                 c = ws.cell(row=curr_row, column=start_c)
            else:
                 ws.merge_cells(start_row=curr_row, start_column=start_c, end_row=curr_row, end_column=end_c)
                 c = ws.cell(row=curr_row, column=start_c)
            
            c.value = h_text
            c.font = self.font_header
            c.fill = PatternFill(start_color=self.azul_primario, end_color=self.azul_primario, fill_type="solid")
            c.alignment = Alignment(horizontal='center', vertical='center')
            
        curr_row += 1
        
        tbody = section.find('tbody')
        if tbody:
            for tr in tbody.find_all('tr'):
                tds = tr.find_all('td')
                if len(tds) < 6: continue
                
                # Item
                ws.cell(row=curr_row, column=2, value=tds[0].get_text(strip=True)).alignment = Alignment(horizontal='center')
                
                # Desc
                ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=6)
                c_desc = ws.cell(row=curr_row, column=3, value=tds[1].get_text(strip=True))
                c_desc.alignment = Alignment(wrap_text=True, vertical='center')
                
                # Cant
                ws.cell(row=curr_row, column=7, value=self.clean_numeric(tds[2].get_text(strip=True)))
                
                # Und
                ws.cell(row=curr_row, column=8, value=tds[3].get_text(strip=True)).alignment = Alignment(horizontal='center')
                
                # PU
                ws.merge_cells(start_row=curr_row, start_column=9, end_row=curr_row, end_column=10)
                c_pu = ws.cell(row=curr_row, column=9, value=self.clean_numeric(tds[4].get_text(strip=True)))
                c_pu.number_format = '"$ "#,##0.00'
                
                # Total
                ws.merge_cells(start_row=curr_row, start_column=11, end_row=curr_row, end_column=12)
                c_tot = ws.cell(row=curr_row, column=11, value=self.clean_numeric(tds[5].get_text(strip=True)))
                c_tot.number_format = '"$ "#,##0.00'
                
                # Borders?
                # Need to set border for all merged cells range
                # Simplifying: Set border on top-left of merge is usually enough if merge handled, but for correctness:
                # We should set border on all edges.
                # For now using simple logic.
                
                curr_row += 1
        return curr_row

    def _render_totales(self, section, ws, start_row):
        curr_row = start_row + 1
        for row in section.find_all(class_='totales-row'):
            lbl = row.find(class_='totales-label').get_text(strip=True)
            val = self.clean_numeric(row.find(class_='totales-valor').get_text(strip=True))
            
            # Label: I-J
            ws.merge_cells(start_row=curr_row, start_column=9, end_row=curr_row, end_column=10)
            c_lbl = ws.cell(row=curr_row, column=9, value=lbl)
            c_lbl.font = Font(bold=True, color=self.azul_secundario)
            c_lbl.alignment = Alignment(horizontal='right')
            
            # Value: K-L
            ws.merge_cells(start_row=curr_row, start_column=11, end_row=curr_row, end_column=12)
            c_val = ws.cell(row=curr_row, column=11, value=val)
            c_val.number_format = '"$ "#,##0.00'
            c_val.font = Font(bold=True)
            
            if "TOTAL:" in lbl:
                c_lbl.fill = PatternFill(start_color=self.azul_primario, end_color=self.azul_primario, fill_type="solid")
                c_lbl.font = Font(bold=True, color=self.blanco)
                c_val.fill = PatternFill(start_color=self.azul_primario, end_color=self.azul_primario, fill_type="solid")
                c_val.font = Font(bold=True, color=self.blanco)
            curr_row += 1
        return curr_row

    def _render_cronograma(self, box, ws, start_row):
        row = start_row
        col = 2
        for fase in box.find_all(class_='fase-box'):
             if col > 12: 
                 col = 2
                 row += 2
                 
             nombre = fase.find(class_='fase-nombre').get_text(strip=True)
             duracion = fase.find(class_='fase-duracion').get_text(strip=True)
             
             # Use 2 cols per fase?
             end_col = min(col+1, 12)
             ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
             c = ws.cell(row=row, column=col)
             c.value = f"{nombre}\n{duracion}"
             c.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
             c.fill = PatternFill(start_color=self.azul_claro, end_color=self.azul_claro, fill_type="solid")
             side = Side(style='thin', color=self.azul_primario)
             c.border = Border(left=side, right=side, top=side, bottom=side)
             
             col += 2
        return row + 2

    def _render_garantias(self, grid, ws, start_row):
        row = start_row
        col = 2
        for card in grid.find_all(class_='garantia-card'):
            texto = card.find(class_='garantia-texto').get_text(strip=True)
            icon = card.find(class_='garantia-icon').get_text(strip=True)
            
            # Grid of 3? (B-D, E-G?, H-J?)
            # Or automatic flow.
            # Using 3 cols per card.
            end_col = min(col+3, 12)
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
            
            c = ws.cell(row=row, column=col)
            c.value = f"{icon} {texto}"
            c.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            self.set_border(c)
            col += 4
            if col > 12:
                col = 2
                row += 1
        return row + 1

    def convert_informe(self, soup, ws):
        self.convert_cotizacion(soup, ws)

    def convert_html_string(self, html_content, output_path):
        soup = BeautifulSoup(html_content, 'html.parser')
        wb = Workbook()
        ws = wb.active
        ws.title = "Documento Tesla"
        ws.sheet_view.showGridLines = False
        
        # Updated Detection Logic for V9 Templates
        is_cotizacion = bool(soup.find(class_='cotizacion-container') or 
                             soup.find(class_='items-table') or 
                             soup.find(class_='tabla-section'))
        
        if is_cotizacion:
            ws.title = "Cotización"
            self.convert_cotizacion(soup, ws)
        else:
            ws.title = "Informe"
            self.convert_informe(soup, ws)
            
        wb.save(output_path)
        logger.info(f"✅ Excel V10 Mirror Generado: {output_path}")
