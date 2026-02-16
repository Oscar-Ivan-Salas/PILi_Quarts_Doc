import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter

def crear_excel_cotizacion(html_path, output_path):
    # Leer el HTML
    with open(html_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')

    # Crear el libro y la hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"

    # Definir estilos
    azul_primario = "0052A3"
    azul_secundario = "1E40AF"
    gris_claro = "F9FAFB"
    blanco = "FFFFFF"

    # Configuración de columnas (A a F)
    column_widths = [10, 50, 12, 12, 15, 15]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = width

    # --- CABECERA ---
    # Logo Placeholder (Simulado con una celda grande)
    ws.merge_cells('A1:B3')
    cell_logo = ws['A1']
    cell_logo.value = "TESLA ELECTRICIDAD"
    cell_logo.fill = PatternFill(start_color=azul_primario, end_color=azul_primario, fill_type="solid")
    cell_logo.font = Font(color=blanco, bold=True, size=16)
    cell_logo.alignment = Alignment(horizontal="center", vertical="center")

    # Info Empresa
    empresa_nombre = soup.find(class_='empresa-nombre').get_text(strip=True)
    ws['F1'] = empresa_nombre
    ws['F1'].font = Font(color=azul_primario, bold=True, size=14)
    ws['F1'].alignment = Alignment(horizontal="right")

    detalles = soup.find(class_='empresa-detalles').get_text("\n", strip=True).split("\n")
    for i, linea in enumerate(detalles):
        ws.cell(row=2+i, column=6, value=linea).font = Font(size=9, color="4B5563")
        ws.cell(row=2+i, column=6).alignment = Alignment(horizontal="right")

    # --- TÍTULO DOCUMENTO ---
    ws.merge_cells('A6:F7')
    titulo_h1 = soup.find('h1').get_text(strip=True)
    ws['A6'] = titulo_h1
    ws['A6'].font = Font(color=azul_primario, bold=True, size=20)
    ws['A6'].fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    ws['A6'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Borde izquierdo para el título (como en el HTML)
    borde_izq = Border(left=Side(style='thick', color=azul_primario))
    ws['A6'].border = borde_izq
    ws['A7'].border = borde_izq

    # --- INFORMACIÓN CLIENTE Y COTIZACIÓN ---
    info_boxes = soup.find_all(class_='info-box')
    
    # Cliente (Izquierda)
    ws['A9'] = "INFORMACIÓN DEL CLIENTE"
    ws['A9'].font = Font(bold=True, color=azul_primario)
    ws.merge_cells('A9:B9')
    
    row_idx = 10
    cliente_p = info_boxes[0].find_all('p')
    for p in cliente_p:
        text = p.get_text(strip=True)
        ws.cell(row=row_idx, column=1, value=text).font = Font(size=10)
        row_idx += 1

    # Detalles Cotización (Derecha)
    ws['E9'] = "DETALLES DE COTIZACIÓN"
    ws['E9'].font = Font(bold=True, color=azul_primario)
    ws.merge_cells('E9:F9')
    
    row_idx = 10
    cotiz_p = info_boxes[1].find_all('p')
    for p in cotiz_p:
        text = p.get_text(strip=True)
        ws.cell(row=row_idx, column=5, value=text).font = Font(size=10)
        ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="right")
        ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=6)
        row_idx += 1

    # --- TABLA DE ITEMS ---
    start_table_row = 16
    ws.cell(row=start_table_row, column=1, value="DETALLE DE LA COTIZACIÓN").font = Font(bold=True, size=12, color=azul_primario)
    start_table_row += 2

    headers = ["ITEM", "DESCRIPCIÓN", "CANT.", "UNIDAD", "P. UNIT.", "TOTAL"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=start_table_row, column=i+1, value=h)
        cell.font = Font(color=blanco, bold=True)
        cell.fill = PatternFill(start_color=azul_primario, end_color=azul_primario, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Extraer filas de la tabla
    rows = soup.find('tbody').find_all('tr')
    current_row = start_table_row + 1
    for r in rows:
        cols = r.find_all('td')
        for i, c in enumerate(cols):
            val = c.get_text(strip=True)
            # Intentar convertir a número si es posible
            clean_val = val.replace('$', '').replace(',', '').strip()
            try:
                if i >= 2: # Columnas numéricas
                    cell_val = float(clean_val)
                else:
                    cell_val = val
            except:
                cell_val = val
            
            cell = ws.cell(row=current_row, column=i+1, value=cell_val)
            cell.alignment = Alignment(horizontal="left" if i < 2 else "right")
            # Estilo cebra
            if current_row % 2 == 0:
                cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            
            # Formato moneda para las últimas dos columnas
            if i >= 4:
                cell.number_format = '"$"#,##0.00'
        current_row += 1

    # --- TOTALES ---
    current_row += 1
    totales_rows = soup.find_all(class_='totales-row')
    for tr in totales_rows:
        label = tr.find(class_='totales-label').get_text(strip=True)
        valor = tr.find(class_='totales-valor').get_text(strip=True)
        
        ws.cell(row=current_row, column=5, value=label).font = Font(bold=True)
        
        # Limpiar valor para Excel
        clean_valor = valor.replace('$', '').replace('{{SUBTOTAL}}', '0').replace('{{IGV}}', '0').replace('{{TOTAL}}', '0').strip()
        try:
            num_valor = float(clean_valor)
        except:
            num_valor = valor

        cell_v = ws.cell(row=current_row, column=6, value=num_valor)
        cell_v.font = Font(bold=True)
        cell_v.number_format = '"$"#,##0.00'
        
        if "TOTAL:" in label:
            ws.cell(row=current_row, column=5).fill = PatternFill(start_color=azul_primario, end_color=azul_primario, fill_type="solid")
            ws.cell(row=current_row, column=5).font = Font(color=blanco, bold=True)
            cell_v.fill = PatternFill(start_color=azul_primario, end_color=azul_primario, fill_type="solid")
            cell_v.font = Font(color=blanco, bold=True)
            
        current_row += 1

    # --- PIE DE PÁGINA ---
    current_row += 3
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    footer_emp = soup.find(class_='footer-empresa').get_text(strip=True)
    ws.cell(row=current_row, column=1, value=footer_emp).font = Font(bold=True, color=azul_primario)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center")
    
    current_row += 1
    footer_contacts = soup.find_all(class_='footer-contacto')
    for fc in footer_contacts:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        ws.cell(row=current_row, column=1, value=fc.get_text(strip=True)).font = Font(size=8, color="6B7280")
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center")
        current_row += 1

    # Guardar el archivo
    wb.save(output_path)
    print(f"Excel generado exitosamente en: {output_path}")

if __name__ == "__main__":
    html_file = "/home/ubuntu/upload/PLANTILLA_HTML_COTIZACION_COMPLEJA.html"
    excel_file = "/home/ubuntu/Cotizacion_Tesla.xlsx"
    crear_excel_cotizacion(html_file, excel_file)
