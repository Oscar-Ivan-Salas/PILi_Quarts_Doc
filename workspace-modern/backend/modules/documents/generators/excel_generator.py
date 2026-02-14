"""
Document Generator Module - Excel Generator
Enterprise-grade Excel generation using openpyxl
Following clean-code and python-patterns skills
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional
from datetime import datetime
from io import BytesIO
import logging

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Professional Excel generator using openpyxl.
    
    Features:
    - Multiple sheets
    - Custom styling
    - Formulas
    - Charts (future)
    - Professional formatting
    
    Following clean-code: Single Responsibility
    """
    
    def __init__(self):
        """Initialize Excel generator"""
        self.header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.title_font = Font(color="1E40AF", bold=True, size=16)
        self.border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        logger.info("Excel Generator initialized")
    
    def generate_cotizacion(
        self,
        data: Dict[str, Any],
        output_path: Optional[str] = None
    ) -> BytesIO:
        """
        Generate professional quotation Excel workbook.
        
        Following python-patterns: Type hints and structure
        
        Args:
            data: Quotation data
            output_path: Optional file path to save workbook
        
        Returns:
            BytesIO buffer with Excel workbook
        """
        logger.info("Generating quotation Excel workbook")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cotización"
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:E{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "COTIZACIÓN"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # Document info
        ws[f'A{row}'] = "No. Cotización:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = data.get('numero', 'N/A')
        row += 1
        
        ws[f'A{row}'] = "Fecha:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = data.get('fecha', datetime.now().strftime('%d/%m/%Y'))
        row += 1
        
        ws[f'A{row}'] = "Válida hasta:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = data.get('valida_hasta', 'N/A')
        row += 2
        
        # Client info section
        ws.merge_cells(f'A{row}:E{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "DATOS DEL CLIENTE"
        section_cell.font = Font(bold=True, size=14, color="3B82F6")
        row += 1
        
        cliente = data.get('cliente', {})
        client_fields = [
            ('Nombre:', cliente.get('nombre', 'N/A')),
            ('Empresa:', cliente.get('empresa', 'N/A')),
            ('Email:', cliente.get('email', 'N/A')),
            ('Teléfono:', cliente.get('telefono', 'N/A')),
            ('Dirección:', cliente.get('direccion', 'N/A')),
        ]
        
        for label, value in client_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Items section
        ws.merge_cells(f'A{row}:E{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "DETALLE DE SERVICIOS"
        section_cell.font = Font(bold=True, size=14, color="3B82F6")
        row += 1
        
        # Table headers
        headers = ['#', 'Descripción', 'Cantidad', 'Precio Unit.', 'Subtotal']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        header_row = row
        row += 1
        
        # Add items
        items = data.get('items', [])
        for idx, item in enumerate(items, 1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=item.get('descripcion', ''))
            ws.cell(row=row, column=3, value=item.get('cantidad', 0))
            ws.cell(row=row, column=4, value=item.get('precio_unitario', 0))
            ws.cell(row=row, column=5, value=item.get('subtotal', 0))
            
            # Format currency columns
            ws.cell(row=row, column=4).number_format = '$#,##0.00'
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            
            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        row += 1
        
        # Totals
        totales = data.get('totales', {})
        
        ws[f'D{row}'] = "Subtotal:"
        ws[f'D{row}'].font = Font(bold=True)
        ws[f'E{row}'] = totales.get('subtotal', 0)
        ws[f'E{row}'].number_format = '$#,##0.00'
        row += 1
        
        ws[f'D{row}'] = "IVA (16%):"
        ws[f'D{row}'].font = Font(bold=True)
        ws[f'E{row}'] = totales.get('iva', 0)
        ws[f'E{row}'].number_format = '$#,##0.00'
        row += 1
        
        ws[f'D{row}'] = "TOTAL:"
        ws[f'D{row}'].font = Font(bold=True, size=14, color="1E40AF")
        ws[f'E{row}'] = totales.get('total', 0)
        ws[f'E{row}'].number_format = '$#,##0.00'
        ws[f'E{row}'].font = Font(bold=True, size=14, color="1E40AF")
        row += 2
        
        # Terms
        if 'terminos' in data:
            ws.merge_cells(f'A{row}:E{row}')
            section_cell = ws[f'A{row}']
            section_cell.value = "TÉRMINOS Y CONDICIONES"
            section_cell.font = Font(bold=True, size=14, color="3B82F6")
            row += 1
            
            ws.merge_cells(f'A{row}:E{row+2}')
            terms_cell = ws[f'A{row}']
            terms_cell.value = data['terminos']
            terms_cell.alignment = Alignment(wrap_text=True, vertical='top')
            row += 3
        
        # Footer
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        footer_cell = ws[f'A{row}']
        footer_cell.value = (
            f"Generado por PILi Quarts - {datetime.now().strftime('%d/%m/%Y %H:%M')} | "
            f"{data.get('empresa_nombre', 'PILi Engineering')}"
        )
        footer_cell.font = Font(size=9, color="6B7280")
        footer_cell.alignment = Alignment(horizontal='center')
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer if not output_path else output_path)
        buffer.seek(0)
        
        logger.info("Quotation Excel workbook generated successfully")
        return buffer
    
    def generate_presupuesto(
        self,
        data: Dict[str, Any],
        output_path: Optional[str] = None
    ) -> BytesIO:
        """
        Generate budget Excel with multiple sheets.
        
        Following python-patterns: Multiple sheets for complex data
        """
        logger.info("Generating budget Excel workbook")
        
        wb = Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        # ... (similar to cotizacion)
        
        # Sheet 2: Detailed breakdown
        ws_detail = wb.create_sheet("Detalle")
        # ... (detailed items)
        
        # Sheet 3: Analysis
        ws_analysis = wb.create_sheet("Análisis")
        # ... (charts and analysis)
        
        buffer = BytesIO()
        wb.save(buffer if not output_path else output_path)
        buffer.seek(0)
        
        logger.info("Budget Excel workbook generated successfully")
        return buffer

    def generate_proyecto_simple(self, data: Dict[str, Any], output_path: Optional[str] = None) -> BytesIO:
        """Generate Simple Project Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Proyecto Simple"
        
        # Header
        ws.merge_cells('A1:B1')
        ws['A1'] = "FICHA DE PROYECTO"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Details
        fields = [
            ("Código", data.get('codigo', 'N/A')),
            ("Nombre", data.get('nombre', 'N/A')),
            ("Cliente", data.get('cliente', {}).get('nombre', 'N/A')),
            ("Duración", data.get('duracion_total', 'N/A')),
            ("Presupuesto", f"${data.get('presupuesto', 0):,.2f}"),
            ("Estado", data.get('estado', 'En Ejecución'))
        ]
        
        for i, (label, value) in enumerate(fields, 3):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'] = value

        # Save
        buffer = BytesIO()
        wb.save(buffer if not output_path else output_path)
        buffer.seek(0)
        return buffer

    def generate_proyecto_complejo(self, data: Dict[str, Any], output_path: Optional[str] = None) -> BytesIO:
        """Generate Complex PMI Project Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Project Charter"
        
        # Header
        ws.merge_cells('A1:D1')
        ws['A1'] = "PROJECT CHARTER (PMI)"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # KPI Section
        ws['A3'] = "KPIs del Proyecto"
        ws['A3'].font = Font(bold=True, size=12, color="3B82F6")
        
        kpis = [
            ("SPI (Cronograma)", data.get('spi', 1.0)),
            ("CPI (Costos)", data.get('cpi', 1.0)),
            ("Valor Ganado (EV)", f"${data.get('ev', 0):,.2f}"),
            ("Valor Planificado (PV)", f"${data.get('pv', 0):,.2f}")
        ]
        
        for i, (label, value) in enumerate(kpis, 4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            
        # Resources Table
        row = 10
        ws[f'A{row}'] = "Recursos Asignados"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="3B82F6")
        row += 1
        
        headers = ["Recurso", "Rol", "Asignación %", "Costo/Hora"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
        
        buffer = BytesIO()
        wb.save(buffer if not output_path else output_path)
        buffer.seek(0)
        return buffer

    def generate_informe_tecnico(self, data: Dict[str, Any], output_path: Optional[str] = None) -> BytesIO:
        """Generate Technical Report Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe Técnico"
        
        ws['A1'] = data.get('titulo', 'INFORME TÉCNICO')
        ws['A1'].font = self.title_font
        
        # Sections as rows
        row = 3
        for section in data.get('secciones', []):
            ws[f'A{row}'] = section.get('titulo', 'Sección')
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            ws[f'A{row}'] = section.get('contenido', '')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 2
            
        buffer = BytesIO()
        wb.save(buffer if not output_path else output_path)
        buffer.seek(0)
        return buffer

    def generate_informe_ejecutivo(self, data: Dict[str, Any], output_path: Optional[str] = None) -> BytesIO:
        """Generate Executive Report Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen Ejecutivo"
        
        ws['A1'] = data.get('titulo', 'INFORME EJECUTIVO')
        ws['A1'].font = self.title_font
        
        # Financial Summary
        row = 3
        ws[f'A{row}'] = "Resumen Financiero"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="3B82F6")
        row += 1
        
        financials = [
            ("ROI Esperado", f"{data.get('roi', 0)}%"),
            ("TIR", f"{data.get('tir', 0)}%"),
            ("Payback", f"{data.get('payback', 0)} meses"),
            ("Ahorro Anual", f"${data.get('ahorro_anual', 0):,.2f}")
        ]
        
        for label, value in financials:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
            
        buffer = BytesIO()
        wb.save(buffer if not output_path else output_path)
        buffer.seek(0)
        return buffer
