
import pandas as pd
import logging
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("TemplateFactory")

# Colores Corporativos
COLOR_PRIMARY = "0052A3"    # Azul Tesla Oscuro
COLOR_SECONDARY = "1E40AF"  # Azul Medio
COLOR_ACCENT = "3B82F6"     # Azul Claro
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_BORDER = "BDC3C7"

def _aplicar_estilos_base(ws):
    ws.sheet_view.showGridLines = False

def _dibujar_header_tesla(ws, titulo_doc):
    """Dibuja el Header Tesla V1 (Placeholder Estático)"""
    # Logo Text
    ws['B2'] = "TESLA"
    ws['B2'].font = Font(name='Arial Black', size=24, bold=True, color=COLOR_PRIMARY)
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['B3'] = "Electricidad y Automatización"
    ws['B3'].font = Font(name='Segoe UI', size=9, bold=True, color="6B7280")
    
    # Doc Title
    ws['D2'] = titulo_doc
    ws['D2'].font = Font(name='Segoe UI', size=18, bold=True, color=COLOR_SECONDARY)
    ws['D2'].alignment = Alignment(horizontal='right', vertical='center')
    
    # Info Empresa
    ws['D3'] = "TESLA ELECTRICIDAD Y AUTOMATIZACIÓN S.A.C."
    ws['D3'].alignment = Alignment(horizontal='right')
    ws['D3'].font = Font(size=9, color="4B5563")
    
    ws['D4'] = "RUC: 20601138787"
    ws['D4'].alignment = Alignment(horizontal='right')
    ws['D4'].font = Font(size=9, color="4B5563")
    
    ws['D5'] = "Email: ingenieria.teslaelectricidad@gmail.com"
    ws['D5'].alignment = Alignment(horizontal='right')
    ws['D5'].font = Font(size=9, color="4B5563")
    
    # Linea
    for col in range(2, 8):
        cell = ws.cell(row=6, column=col)
        cell.border = Border(bottom=Side(style='medium', color=COLOR_PRIMARY))

def create_template_cotizacion(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"
    _aplicar_estilos_base(ws)
    _dibujar_header_tesla(ws, "COTIZACIÓN DE SERVICIOS")
    
    # Placeholder Info Grid (Rows 8-12)
    ws['B8'] = "INFORMACIÓN DEL CLIENTE"
    ws['B8'].font = Font(bold=True, color=COLOR_PRIMARY)
    
    labels = ["Cliente:", "Proyecto:", "Fecha:", "Validez:"]
    for i, lab in enumerate(labels):
        ws[f'B{9+i}'] = lab
        ws[f'C{9+i}'] = "{{ " + lab.replace(":","").lower() + " }}" # Jinja-like placeholder
        
    # Headers Table (Row 14)
    headers = ['Item', 'Descripción', 'Cant', 'Und', 'P.Unit', 'Total']
    for idx, h in enumerate(headers):
        cell = ws.cell(row=14, column=idx+2)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        
    # Column Widths
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    wb.save(path)
    logger.info(f"✅ Created Template: {path}")

def create_template_proyecto(path):
    wb = Workbook()
    
    # Sheet 1: Charter
    ws1 = wb.active
    ws1.title = "Project Charter"
    _aplicar_estilos_base(ws1)
    _dibujar_header_tesla(ws1, "PROJECT CHARTER")
    
    # Sheet 2: Cronograma
    ws2 = wb.create_sheet("Cronograma")
    _aplicar_estilos_base(ws2)
    _dibujar_header_tesla(ws2, "CRONOGRAMA GANTT")
    
    # Sheet 3: Riesgos
    ws3 = wb.create_sheet("Riesgos")
    _aplicar_estilos_base(ws3)
    _dibujar_header_tesla(ws3, "MATRIZ DE RIESGOS")
    
    wb.save(path)
    logger.info(f"✅ Created Template: {path}")

def create_template_informe(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe"
    _aplicar_estilos_base(ws)
    _dibujar_header_tesla(ws, "INFORME TÉCNICO")
    
    wb.save(path)
    logger.info(f"✅ Created Template: {path}")

if __name__ == "__main__":
    base_dir = r"e:\PILi_Quarts\workspace-modern\backend\app\templates\excel"
    os.makedirs(base_dir, exist_ok=True)
    
    create_template_cotizacion(os.path.join(base_dir, "PLANTILLA_COTIZACION.xlsx"))
    create_template_proyecto(os.path.join(base_dir, "PLANTILLA_PROYECTO.xlsx"))
    create_template_informe(os.path.join(base_dir, "PLANTILLA_INFORME.xlsx"))
