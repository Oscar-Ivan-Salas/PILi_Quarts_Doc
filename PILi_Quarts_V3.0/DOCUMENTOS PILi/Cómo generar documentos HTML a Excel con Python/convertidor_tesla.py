import os
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class TeslaExcelConverter:
    def __init__(self, azul_primario="0052A3", azul_secundario="1E40AF", gris_claro="F9FAFB"):
        self.azul_primario = azul_primario
        self.azul_secundario = azul_secundario
        self.gris_claro = gris_claro
        self.blanco = "FFFFFF"

    def clean_numeric(self, text):
        if not text: return 0.0
        clean = re.sub(r'[^\d.]', '', text.replace(',', ''))
        try:
            return float(clean)
        except ValueError:
            return 0.0

    def apply_header_style(self, cell):
        cell.font = Font(color=self.blanco, bold=True)
        cell.fill = PatternFill(start_color=self.azul_primario, end_color=self.azul_primario, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    def convert_cotizacion(self, soup, ws):
        column_widths = [10, 50, 12, 12, 15, 15]
        for i, width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = width

        ws.merge_cells('A1:B3')
        ws['A1'] = "TESLA ELECTRICIDAD"
        ws['A1'].fill = PatternFill(start_color=self.azul_primario, end_color=self.azul_primario, fill_type="solid")
        ws['A1'].font = Font(color=self.blanco, bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

        emp_nombre = soup.find(class_='empresa-nombre')
        if emp_nombre:
            ws['F1'] = emp_nombre.get_text(strip=True)
            ws['F1'].font = Font(color=self.azul_primario, bold=True, size=14)
            ws['F1'].alignment = Alignment(horizontal="right")

        ws.merge_cells('A6:F7')
        h1 = soup.find('h1')
        if h1:
            ws['A6'] = h1.get_text(strip=True)
            ws['A6'].font = Font(color=self.azul_primario, bold=True, size=20)
            ws['A6'].fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
            ws['A6'].alignment = Alignment(horizontal="center", vertical="center")

        tbody = soup.find('tbody')
        if tbody:
            start_row = 18
            headers = ["ITEM", "DESCRIPCIÓN", "CANT.", "UNIDAD", "P. UNIT.", "TOTAL"]
            for i, h in enumerate(headers):
                self.apply_header_style(ws.cell(row=start_row, column=i+1, value=h))
            
            curr_row = start_row + 1
            for tr in tbody.find_all('tr'):
                for i, td in enumerate(tr.find_all('td')):
                    val = td.get_text(strip=True)
                    cell = ws.cell(row=curr_row, column=i+1, value=self.clean_numeric(val) if i >= 2 and i != 3 else val)
                    if i >= 4: cell.number_format = '"$"#,##0.00'
                curr_row += 1

    def convert_informe_narrativo(self, soup, ws):
        """Maneja tanto Informe Ejecutivo como Informe Técnico."""
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 90
        
        curr_row = 1
        
        # 1. Portada (APA o Técnica)
        portada = soup.find(class_=['portada-apa', 'portada'])
        if portada:
            ws.merge_cells('B2:B4')
            title = portada.find('h1').get_text(strip=True) if portada.find('h1') else "INFORME"
            ws['B2'] = title
            ws['B2'].font = Font(bold=True, size=18, color=self.azul_primario)
            ws['B2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            sub = portada.find('h2')
            if sub:
                ws.cell(row=5, column=2, value=sub.get_text(strip=True)).font = Font(bold=True, size=14)
                ws.cell(row=5, column=2).alignment = Alignment(horizontal="center")
            
            curr_row = 7
            portada_info = portada.find(class_=['portada-info', 'portada-detalles'])
            if portada_info:
                for div in portada_info.find_all('div'):
                    ws.cell(row=curr_row, column=2, value=div.get_text(strip=True)).alignment = Alignment(horizontal="center")
                    curr_row += 1
        
        curr_row += 3
        
        # 2. Metadatos (si existen)
        metadatos = soup.find(class_='metadatos')
        if metadatos:
            ws.cell(row=curr_row, column=2, value="METADATOS DEL INFORME").font = Font(bold=True, color=self.azul_primario)
            curr_row += 1
            for item in metadatos.find_all(class_='metadatos-item'):
                label = item.find(class_='metadatos-label').get_text(strip=True)
                val = item.find(class_='metadatos-value').get_text(strip=True)
                ws.cell(row=curr_row, column=2, value=f"{label} {val}")
                curr_row += 1
            curr_row += 2

        # 3. Resumen Ejecutivo
        summary = soup.find(class_=['executive-summary', 'resumen-ejecutivo'])
        if summary:
            title_text = summary.find(['h2', 'h3']).get_text(strip=True) if summary.find(['h2', 'h3']) else "RESUMEN"
            ws.cell(row=curr_row, column=2, value=title_text).font = Font(bold=True, color=self.azul_primario, size=14)
            ws.cell(row=curr_row, column=2).fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
            curr_row += 2
            for p in summary.find_all('p'):
                ws.cell(row=curr_row, column=2, value=p.get_text(strip=True)).alignment = Alignment(wrap_text=True)
                curr_row += 1
            curr_row += 2

        # 4. Secciones
        secciones = soup.find_all(class_='seccion')
        for sec in secciones:
            h2 = sec.find('h2')
            if h2:
                ws.cell(row=curr_row, column=2, value=h2.get_text(strip=True)).font = Font(bold=True, color=self.azul_primario, size=12)
                curr_row += 1
            
            for elem in sec.find_all(['p', 'h3', 'li', 'h4']):
                text = elem.get_text(strip=True)
                prefix = "• " if elem.name == 'li' else ""
                cell = ws.cell(row=curr_row, column=2, value=f"{prefix}{text}")
                if elem.name in ['h3', 'h4']: 
                    cell.font = Font(bold=True, color=self.azul_secundario)
                cell.alignment = Alignment(wrap_text=True)
                curr_row += 1
            curr_row += 1

        # 5. Tablas Técnicas
        tables = soup.find_all('table')
        for table in tables:
            curr_row += 1
            headers = [th.get_text(strip=True) for th in table.find_all('th')]
            if headers:
                for i, h in enumerate(headers):
                    cell = ws.cell(row=curr_row, column=i+2, value=h)
                    cell.font = Font(bold=True, color=self.blanco)
                    cell.fill = PatternFill(start_color=self.azul_primario, end_color=self.azul_primario, fill_type="solid")
                curr_row += 1
            
            for tr in table.find_all('tr'):
                tds = tr.find_all('td')
                if not tds: continue
                for i, td in enumerate(tds):
                    ws.cell(row=curr_row, column=i+2, value=td.get_text(strip=True))
                curr_row += 1
            curr_row += 2

    def convert(self, html_path, output_path):
        with open(html_path, 'r', encoding='utf-8') as f:
            soup = BeautifulSoup(f, 'html.parser')

        wb = Workbook()
        ws = wb.active
        
        # Lógica de detección mejorada
        if soup.find(class_=['portada-apa', 'portada', 'seccion']):
            ws.title = "Informe"
            self.convert_informe_narrativo(soup, ws)
        else:
            ws.title = "Cotización"
            self.convert_cotizacion(soup, ws)

        wb.save(output_path)
        print(f"Excel generado: {output_path}")

if __name__ == "__main__":
    converter = TeslaExcelConverter()
    templates = [
        ("/home/ubuntu/upload/PLANTILLA_HTML_COTIZACION_COMPLEJA.html", "/home/ubuntu/Cotizacion_Tesla_Compleja.xlsx"),
        ("/home/ubuntu/upload/PLANTILLA_HTML_COTIZACION_SIMPLE.html", "/home/ubuntu/Cotizacion_Tesla_Simple.xlsx"),
        ("/home/ubuntu/upload/PLANTILLA_HTML_INFORME_EJECUTIVO_APA.html", "/home/ubuntu/Informe_Ejecutivo_Tesla.xlsx"),
        ("/home/ubuntu/upload/PLANTILLA_HTML_INFORME_TECNICO.html", "/home/ubuntu/Informe_Tecnico_Tesla.xlsx")
    ]
    for html, excel in templates:
        if os.path.exists(html):
            converter.convert(html, excel)
