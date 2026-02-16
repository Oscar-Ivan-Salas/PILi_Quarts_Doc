
import logging
import io
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage

# N04 INTERNAL IMPORTS ONLY
# No app.services or legacy imports allowed here.

logger = logging.getLogger("N04_ExcelGenerator")

# ══════════════════════════════════════════════════════════════════════════════
# 🏭 EXCEL GENERATION STRATEGY (V10 ZONING ENGINE)
# ══════════════════════════════════════════════════════════════════════════════

def generate_excel_document(mode: str, items: list, totals: dict, branding: dict, metadata: dict) -> dict:
    """
    Main entry point for Excel generation (V10 Standard).
    """
    try:
        # 1. Create Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cotizacion" # Default
        
        # 2. Setup Styles
        _setup_styles(wb)
        
        # 3. Add Branding (Logo - Zone 1)
        if branding.get("logo"):
             _insert_logo(ws, branding["logo"])
        
        # 4. Generate Content based on Mode
        title_map = {
            "cotizacion_simple": "COTIZACIÓN DE SERVICIOS",
            "cotizacion_compleja": "COTIZACIÓN TÉCNICA DETALLADA",
            "proyecto_simple": "PRESUPUESTO DE PROYECTO",
            "proyecto_complejo": "PRESUPUESTO GENERAL DE OBRA",
            "informe_tecnico": "INFORME TÉCNICO",
            "informe_ejecutivo": "INFORME EJECUTIVO GERENCIAL"
        }
        
        doc_title = title_map.get(mode, "DOCUMENTO TÉCNICO")
        
        # ZONE 1 & 2: Header Info & Client Data
        current_row = _write_v10_header(ws, doc_title, metadata, branding["color"])
        
        # ZONE 3: Body (Items)
        current_row = _write_v10_items_table(ws, items, current_row, branding["color"])
        
        # ZONE 4: Totals & Footer
        _write_v10_totals(ws, totals, current_row, branding["color"], metadata)
        
        # 5. Output
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        b64_data = base64.b64encode(output.read()).decode('utf-8')
        
        # STRICT NAMING CONVENTION: INFORME_TECNICO_0001_TESLA.XLSX
        # User requirement: Todos los archivos en MAYÚSCULAS: INFORME_TECNICO_0001_TESLA.PDF
        doc_label = str(mode).upper().replace(" ", "_").replace("-", "_")
        svc_id_fmt = str(metadata.get('service_id', '0001')).zfill(4)
        final_name = f"{doc_label}_{svc_id_fmt}_TESLA.xlsx"
        
        return {
            "success": True,
            "filename": final_name,
            "file_b64": b64_data,
            "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }

    except Exception as e:
        logger.error(f"Excel Generation Error: {e}", exc_info=True)
        raise e

# ══════════════════════════════════════════════════════════════════════════════
# 🎨 HELPER FUNCTIONS (V10 Standard)
# ══════════════════════════════════════════════════════════════════════════════

def _setup_styles(wb):
    # Basic setup, could be expanded
    pass

def _insert_logo(ws, logo_b64):
    try:
        if not logo_b64: return
        img_data = base64.b64decode(logo_b64)
        img = PILImage.open(io.BytesIO(img_data))
        # Resize logic if needed
        out_img = io.BytesIO()
        img.save(out_img, format="PNG")
        out_img.seek(0)
        
        xl_img = OpenpyxlImage(out_img)
        xl_img.anchor = 'A1' 
        # Resize roughly to fit A1:C4
        xl_img.width = 180
        xl_img.height = 60
        ws.add_image(xl_img)
    except Exception as e:
        logger.warning(f"Logo insertion failed: {e}")

def _write_v10_header(ws, title, metadata, color):
    # Dimensions
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40 # Description
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Title (Centered D2:G2)
    ws.merge_cells('D2:G2')
    title_cell = ws['D2']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, color=color.replace("#", ""))
    title_cell.alignment = Alignment(horizontal='center')

    # Company Info (D3:G3)
    ws.merge_cells('D3:G3')
    ws['D3'] = "TESLA ELECTRICIDAD Y AUTOMATIZACIÓN S.A.C."
    ws['D3'].font = Font(bold=True)
    ws['D3'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('D4:G4')
    ws['D4'] = "RUC: 20601138787"
    ws['D4'].alignment = Alignment(horizontal='center')

    # Client Box (Row 6-10)
    # Box Border Style
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # Client Data Extraction
    client_name = metadata.get('client_info', {}).get('nombre', 'N/A')
    client_ruc = metadata.get('client_info', {}).get('ruc', 'N/A')
    client_addr = metadata.get('client_info', {}).get('direccion', 'N/A')
    client_date = metadata.get('client_info', {}).get('fecha', datetime.now().strftime('%d/%m/%Y'))
    
    # Zone 2: Metadatos (Filas 10-15)
    start_row = 10
    
    # Row 10: Client Name
    ws['B10'] = "SEÑORES:"
    ws['B10'].font = Font(bold=True)
    ws['C10'] = client_name
    
    # Row 11: RUC
    ws['B11'] = "RUC:"
    ws['B11'].font = Font(bold=True)
    ws['C11'] = client_ruc
    
    # Row 12: Address
    ws['B12'] = "DIRECCIÓN:"
    ws['B12'].font = Font(bold=True)
    ws['C12'] = client_addr
    
    # Row 10 (Right Side): Date / Quote No
    ws['F10'] = "FECHA:"
    ws['F10'].font = Font(bold=True)
    ws['G10'] = client_date
    
    ws['F11'] = "COTIZACIÓN Nº:"
    ws['F11'].font = Font(bold=True)
    ws['G11'] = metadata.get('codigo', 'COT-XXXX')

    ws['F12'] = "MONEDA:"
    ws['F12'].font = Font(bold=True)
    ws['G12'] = metadata.get('MONEDA', 'USD')

    return 18 # Start Content at Row 18

def _write_v10_items_table(ws, items, start_row, color):
    # Headers
    headers = ["ITEM", "DESCRIPCIÓN", "CANT.", "UNIDAD", "P. UNIT.", "TOTAL"]
    # Mapping to columns: A=1, B=2 (Item), C=3 (Desc), D=4 (Cant), E=5 (Unit), F=6 (Price), G=7 (Total)
    # Let's align with V10 layout:
    # B: Item (Index)
    # C: Description
    # D: Cant
    # E: Unit
    # F: Price
    # G: Total
    
    header_row = start_row
    
    col_map = {
        2: "ITEM",
        3: "DESCRIPCIÓN",
        4: "CANT.",
        5: "UNIDAD",
        6: "P. UNIT.",
        7: "TOTAL"
    }
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col_idx, text in col_map.items():
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color.replace("#", ""), end_color=color.replace("#", ""), fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    current_row = header_row + 1
    for idx, item in enumerate(items, 1):
        # B: Item
        c = ws.cell(row=current_row, column=2, value=idx) # 01, 02...
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')
        
        # C: Desc
        c = ws.cell(row=current_row, column=3, value=item.get("descripcion", ""))
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True)
        
        # D: Cant
        c = ws.cell(row=current_row, column=4, value=float(item.get("cantidad", 0)))
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')
        
        # E: Unit
        c = ws.cell(row=current_row, column=5, value=item.get("unidad", "und"))
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')
        
        # F: Price
        c = ws.cell(row=current_row, column=6, value=float(item.get("precio", 0)))
        c.border = thin_border
        c.number_format = '#,##0.00'
        
        # G: Total
        c = ws.cell(row=current_row, column=7, value=float(item.get("total", 0)))
        c.border = thin_border
        c.number_format = '#,##0.00'
        
        current_row += 1
        
    return current_row + 2 # Spacing

def _write_v10_totals(ws, totals, start_row, color, metadata):
    r = start_row
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Subtotal
    ws.cell(row=r, column=6, value="SUBTOTAL:").alignment = Alignment(horizontal='right')
    ws.cell(row=r, column=6).font = Font(bold=True)
    c = ws.cell(row=r, column=7, value=float(totals.get("subtotal", 0)))
    c.number_format = '#,##0.00'
    c.border = thin_border
    
    r += 1
    # IGV
    ws.cell(row=r, column=6, value="IGV (18%):").alignment = Alignment(horizontal='right')
    ws.cell(row=r, column=6).font = Font(bold=True)
    c = ws.cell(row=r, column=7, value=float(totals.get("igv", 0)))
    c.number_format = '#,##0.00'
    c.border = thin_border
    
    r += 1
    # Total
    ws.cell(row=r, column=6, value="TOTAL:").alignment = Alignment(horizontal='right')
    ws.cell(row=r, column=6).font = Font(bold=True)
    c = ws.cell(row=r, column=7, value=float(totals.get("total", 0)))
    c.font = Font(bold=True)
    c.number_format = '#,##0.00'
    c.border = thin_border
    
    # Footer Text (Conditions/Notes) could go here (Zone 4)
    r += 3
    ws['B'+str(r)] = "CONDICIONES DE PAGO:"
    ws['B'+str(r)].font = Font(bold=True, color=color.replace("#", ""))
    
    r += 1
    ws['B'+str(r)] = "• 50% Adelanto, Saldo contra entrega."
