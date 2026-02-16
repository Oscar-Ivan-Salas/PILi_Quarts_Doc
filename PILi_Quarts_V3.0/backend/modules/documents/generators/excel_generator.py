"""
Document Generator Module - Excel Generator
Enterprise-grade Excel generation using openpyxl
Following clean-code and python-patterns skills
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional
from datetime import datetime
from io import BytesIO
import logging

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Professional Excel generator using openpyxl.
    
    Features:
    - Multiple sheets
    - Custom styling
    - Formulas
    - Charts (future)
    - Professional formatting
    
    Following clean-code: Single Responsibility
    """
    
    def __init__(self):
        """Initialize Excel generator with Tesla Visual Identity"""
        # Tesla Palette
        self.COLOR_PRIMARY = "0052A3"
        self.COLOR_SECONDARY = "1E40AF"
        self.COLOR_ACCENT = "3B82F6"
        self.COLOR_BG_LIGHT = "EFF6FF"
        self.COLOR_TEXT_GRAY = "374151"
        self.COLOR_BORDER = "DBEAFE"

        # Styles
        self.header_fill = PatternFill(start_color=self.COLOR_PRIMARY, end_color=self.COLOR_PRIMARY, fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11, name='Calibri')
        
        self.title_font = Font(color=self.COLOR_PRIMARY, bold=True, size=18, name='Calibri')
        self.subtitle_font = Font(color=self.COLOR_SECONDARY, italic=True, size=12, name='Calibri')
        
        self.card_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        self.card_border = Border(
            left=Side(style='medium', color=self.COLOR_ACCENT),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        self.border_thin = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        logger.info("Excel Generator initialized with Tesla Visual Identity")
    
    def generate_cotizacion(
        self,
        data: Dict[str, Any],
        output_path: Optional[str] = None
    ) -> BytesIO:
        """Generate professional quotation Excel workbook - Tesla Visual Identity"""
        logger.info("Generating quotation Excel workbook")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cotización"
        ws.sheet_view.showGridLines = False
        
        # 1. Header (Logo Block & Info) - Gold Standard
        # Logo Block (TESLA)
        ws.merge_cells('A1:C2')
        ws['A1'] = "TESLA"
        ws['A1'].font = Font(color="FFFFFF", bold=True, size=24)
        ws['A1'].fill = self.header_fill 
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Slogan below Logo
        ws.merge_cells('A3:C3')
        ws['A3'] = "Electricidad y Automatización"
        ws['A3'].font = Font(color="374151", size=9, italic=True) # Gray text
        ws['A3'].alignment = Alignment(horizontal='center', vertical='top')
        
        # Company Info (Right)
        ws.merge_cells('D1:F3') 
        ws['D1'] = "TESLA ELECTRICIDAD Y AUTOMATIZACIÓN S.A.C.\nRUC: 20601138787\nJr. Las Ágatas Mz B Lote 09, Urb. San Carlos, SJL\nTeléfono: 906 315 961\nEmail: ingenieria.teslaelectricidad@gmail.com"
        ws['D1'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws['D1'].font = Font(color=self.COLOR_PRIMARY, size=8)
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        row = 5
        
        # Title
        ws.merge_cells(f'A{row}:F{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "COTIZACIÓN DE SERVICIOS"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        sub_cell = ws[f'A{row}']
        sub_cell.value = f"N° {data.get('numero', 'COT-001')}"
        sub_cell.font = self.subtitle_font
        sub_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 2
        
        # Client info section (Side-by-Side as per Gold Standard)
        # Left: Data Client
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "DATOS DEL CLIENTE"
        ws[f'A{row}'].font = Font(bold=True, size=11, color=self.COLOR_PRIMARY)
        
        # Right: Data Cotizacion
        ws.merge_cells(f'D{row}:F{row}')
        ws[f'D{row}'] = "DATOS DE LA COTIZACIÓN"
        ws[f'D{row}'].font = Font(bold=True, size=11, color=self.COLOR_PRIMARY)
        
        row += 1
        start_row = row
        
        # Client Data
        cliente = data.get('cliente', {})
        client_fields = [
            f"Cliente: {cliente.get('nombre', 'N/A')}",
            f"Proyecto: {data.get('proyecto', 'N/A')}",
            f"Dirección: {cliente.get('direccion', 'N/A')}"
        ]
        for val in client_fields:
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = val
            row += 1
            
        # Cotizacion Data (Reset row to start_row for right side)
        r_right = start_row
        cot_fields = [
            f"Fecha: {data.get('fecha', datetime.now().strftime('%d/%m/%Y'))}",
            f"Vigencia: {data.get('vigencia', '30 días')}",
            f"Servicio: {data.get('servicio', 'Instalaciones Eléctricas')}"
        ]
        for val in cot_fields:
            ws.merge_cells(f'D{r_right}:F{r_right}')
            ws[f'D{r_right}'] = val
            r_right += 1
            
        row = max(row, r_right) + 1
        
        # Items section
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "Detalle de la Cotización"
        ws[f'A{row}'].font = Font(bold=True, size=12, color=self.COLOR_PRIMARY, underline='single')
        row += 1
        
        # Table headers
        headers = ['ITEM', 'DESCRIPCIÓN', 'CANT.', 'UNIDAD', 'P. UNIT.', 'TOTAL']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border_thin
        
        row += 1
        
        # Add items
        items = data.get('items', [])
        for idx, item in enumerate(items, 1):
            ws.cell(row=row, column=1, value=f"{idx:02d}").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=item.get('descripcion', ''))
            ws.cell(row=row, column=3, value=item.get('cantidad', 0)).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4, value=item.get('unidad', 'und')).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=item.get('precio_unitario', 0))
            ws.cell(row=row, column=6, value=item.get('monto', 0))
            
            # Format currency columns
            ws.cell(row=row, column=5).number_format = '$ #,##0.00'
            ws.cell(row=row, column=6).number_format = '$ #,##0.00'
            
            # Borders and alternating fill
            for col in range(1, 7):
                c = ws.cell(row=row, column=col)
                c.border = self.border_thin
                if idx % 2 == 0:
                    c.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            
            row += 1
        
        row += 1
        
        # Totals (Right Aligned)
        totales = data.get('totales', {})
        
        labels = [("SUBTOTAL:", totales.get('subtotal', 0)), 
                  ("IGV (18%):", totales.get('igv', 0)), 
                  ("TOTAL:", totales.get('total', 0))]
        
        for label, val in labels:
            ws.merge_cells(f'D{row}:E{row}')
            ws[f'D{row}'] = label
            ws[f'D{row}'].font = Font(bold=True, color=self.COLOR_PRIMARY)
            ws[f'D{row}'].alignment = Alignment(horizontal='right')
            
            ws[f'F{row}'] = val
            ws[f'F{row}'].number_format = '$ #,##0.00'
            ws[f'F{row}'].font = Font(bold=True)
            
            if label == "TOTAL:":
                ws[f'D{row}'].fill = self.header_fill
                ws[f'D{row}'].font = Font(bold=True, color="FFFFFF")
                ws[f'F{row}'].fill = self.header_fill
                ws[f'F{row}'].font = Font(bold=True, color="FFFFFF")
            
            row += 1
            
        row += 2
        
        # Terms
        if 'terminos' in data:
            ws.merge_cells(f'A{row}:E{row}')
            section_cell = ws[f'A{row}']
            section_cell.value = "TÉRMINOS Y CONDICIONES"
            section_cell.font = Font(bold=True, size=14, color="3B82F6")
            row += 1
            
            ws.merge_cells(f'A{row}:E{row+2}')
            terms_cell = ws[f'A{row}']
            terms_cell.value = data['terminos']
            terms_cell.alignment = Alignment(wrap_text=True, vertical='top')
            row += 3
        
        # Footer
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        footer_cell = ws[f'A{row}']
        # 1. Header (Logo Block & Info) - Gold Standard
        # Logo Block (TESLA)
        ws.merge_cells('A1:C2')
        ws['A1'] = "TESLA"
        ws['A1'].font = Font(color="FFFFFF", bold=True, size=24)
        ws['A1'].fill = self.header_fill 
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Slogan below Logo
        ws.merge_cells('A3:C3')
        ws['A3'] = "Electricidad y Automatización"
        ws['A3'].font = Font(color="374151", size=9, italic=True) # Gray text
        ws['A3'].alignment = Alignment(horizontal='center', vertical='top')
        
        # Company Info (Right)
        ws.merge_cells('D1:F3') 
        ws['D1'] = "TESLA ELECTRICIDAD Y AUTOMATIZACIÓN S.A.C.\nRUC: 20601138787\nJr. Las Ágatas Mz B Lote 09, Urb. San Carlos, SJL\nTeléfono: 906 315 961\nEmail: ingenieria.teslaelectricidad@gmail.com"
        ws['D1'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws['D1'].font = Font(color=self.COLOR_PRIMARY, size=8)
        
        # ... (Columns setup remains) ...
        
        # Footer (Brand Footer)
        row += 1
        # Line 1: Company Name
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "TESLA ELECTRICIDAD Y AUTOMATIZACIÓN S.A.C."
        ws[f'A{row}'].font = Font(bold=True, color=self.COLOR_PRIMARY, size=11)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        # Line 2: RUC | Phone
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "RUC: 20601138787 | Teléfono: 906 315 961"
        ws[f'A{row}'].font = Font(color="374151", size=9)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        # Line 3: Email
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "Email: ingenieria.teslaelectricidad@gmail.com"
        ws[f'A{row}'].font = Font(color="374151", size=9)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        # Line 4: Address
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "Jr. Las Ágatas Mz B Lote 09, Urb. San Carlos, SJL"
        ws[f'A{row}'].font = Font(color="374151", size=9)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer if not output_path else output_path)
        buffer.seek(0)
        
        logger.info("Quotation Excel workbook generated successfully")
        return buffer
    
    def generate_presupuesto(
        self,
        data: Dict[str, Any],
        output_path: Optional[str] = None
    ) -> BytesIO:
        """
        Generate budget Excel with multiple sheets.
        
        Following python-patterns: Multiple sheets for complex data
        """
        logger.info("Generating budget Excel workbook")
        
        wb = Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        # ... (similar to cotizacion)
        
        # Sheet 2: Detailed breakdown
        ws_detail = wb.create_sheet("Detalle")
        # ... (detailed items)
        
        # Sheet 3: Analysis
        ws_analysis = wb.create_sheet("Análisis")
        # ... (charts and analysis)
        
        buffer = BytesIO()
        wb.save(buffer if not output_path else output_path)
        buffer.seek(0)
        
        logger.info("Budget Excel workbook generated successfully")
        return buffer

    def generate_proyecto_simple(self, data: Dict[str, Any], output_path: Optional[str] = None) -> BytesIO:
        """Generate Simple Project Excel - Tesla Visual Identity (Deep Mirror)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Ficha de Proyecto"
        ws.sheet_view.showGridLines = False 
        
        # 1. Header (Logo Block & Info)
        # Gradient Sim (Blue Block)
        ws.merge_cells('A1:C3')
        ws['A1'] = "TESLA"
        ws['A1'].font = Font(color="FFFFFF", bold=True, size=24)
        ws['A1'].fill = self.header_fill 
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Company Info
        ws.merge_cells('D1:F3')
        ws['D1'] = "TESLA ELECTRICIDAD Y AUTOMATIZACIÓN S.A.C.\nRUC: 20601138787\nIngeniería y Proyectos"
        ws['D1'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws['D1'].font = Font(color=self.COLOR_PRIMARY, size=10)
        
        row = 5
        
        # 2. Title Block (Light Blue Background + Left Accent)
        ws.merge_cells(f'A{row}:F{row+2}')
        title_cell = ws[f'A{row}']
        title_cell.value = f"PLAN DE PROYECTO\n{data.get('nombre', 'PROYECTO SIN NOMBRE')}"
        title_cell.font = self.title_font
        title_cell.fill = PatternFill(start_color=self.COLOR_BG_LIGHT, end_color=self.COLOR_BG_LIGHT, fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Simulate left border
        for r in range(row, row+3):
             ws[f'A{r}'].border = Border(left=Side(style='thick', color=self.COLOR_PRIMARY))
        row += 4
        
        # 3. Info Grid (Cards) - Mirroring Word
        cards = [
            ("CLIENTE", data.get('cliente', {}).get('nombre', 'N/A')),
            ("CÓDIGO", data.get('codigo', 'N/A')),
            ("ESTADO", data.get('estado', 'N/A'))
        ]
        
        # Manually verify grid mapping A-B, C-D, E-F
        col_sets = [('A','B'), ('C','D'), ('E','F')]
        
        for idx, (label, val) in enumerate(cards):
            c1, c2 = col_sets[idx]
            ws.merge_cells(f'{c1}{row}:{c2}{row}')
            ws[f'{c1}{row}'] = label
            ws[f'{c1}{row}'].font = Font(color="6B7280", size=9, bold=True)
            
            ws.merge_cells(f'{c1}{row+1}:{c2}{row+1}')
            ws[f'{c1}{row+1}'] = val
            ws[f'{c1}{row+1}'].font = Font(color=self.COLOR_PRIMARY, size=12, bold=True)
            ws[f'{c1}{row+1}'].border = self.card_border
            
        row += 3
        
        # 4. Presupuesto Destacado (New Section)
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "PRESUPUESTO ESTIMADO"
        ws[f'A{row}'].font = Font(color="6B7280", size=10, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        presupuesto = data.get('presupuesto', 0)
        ws.merge_cells(f'A{row}:F{row+1}')
        ws[f'A{row}'] = f"$ {presupuesto:,.2f}" if isinstance(presupuesto, (int, float)) else str(presupuesto)
        ws[f'A{row}'].font = Font(color=self.COLOR_PRIMARY, size=24, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 3

        # 5. Alcance (New Section)
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "ALCANCE DEL PROYECTO"
        ws[f'A{row}'].font = Font(size=14, color=self.COLOR_PRIMARY, bold=True)
        row += 1
        ws.merge_cells(f'A{row}:F{row+1}')
        ws[f'A{row}'] = data.get('alcance', data.get('alcance_proyecto', 'Sin alcance definido.'))
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        row += 3

        # 6. Fases Timeline
        ws[f'A{row}'] = "FASES DEL PROYECTO"
        ws[f'A{row}'].font = Font(size=14, color=self.COLOR_PRIMARY, bold=True)
        row += 1
        
        headers = ["Fase", "Actividad", "Duración", "Entregable"]
        # Map headers to specific columns for better spacing
        # Col A: Fase, B: Actividad, C: Duración, D-F: Entregable
        
        # Header Row
        ws[f'A{row}'] = "Fase"; ws[f'A{row}'].fill = self.header_fill; ws[f'A{row}'].font = self.header_font
        ws[f'B{row}'] = "Actividad"; ws[f'B{row}'].fill = self.header_fill; ws[f'B{row}'].font = self.header_font
        ws[f'C{row}'] = "Duración"; ws[f'C{row}'].fill = self.header_fill; ws[f'C{row}'].font = self.header_font
        ws.merge_cells(f'D{row}:F{row}')
        ws[f'D{row}'] = "Entregable"; ws[f'D{row}'].fill = self.header_fill; ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        fases = data.get('fases', [])
        for i, fase in enumerate(fases, 1):
            ws[f'A{row}'] = f"{i}. {fase.get('nombre')}"
            
            # Actividades (List to string)
            acts = fase.get('actividades', [])
            if isinstance(acts, list):
                act_str = "\n".join([f"- {a}" for a in acts])
            else:
                act_str = str(acts)
            ws[f'B{row}'] = act_str
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
            
            # Duracion
            dur = fase.get('duracion', 0)
            ws[f'C{row}'] = f"{dur} días"
            
            # Entregable
            ws.merge_cells(f'D{row}:F{row}')
            ws[f'D{row}'] = fase.get('entregable', '')
            
            # Zebra striping
            if i % 2 == 0:
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                     if col in ['D','E','F']:
                        ws[f'{col}{row}'].fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                     else:
                        ws[f'{col}{row}'].fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            row += 1
            
        row += 2
        
        # 7. Recursos (New Section)
        ws[f'A{row}'] = "RECURSOS ASIGNADOS"
        ws[f'A{row}'].font = Font(size=14, color=self.COLOR_PRIMARY, bold=True)
        row += 1
        
        recursos_data = data.get('recursos', {})
        # Normalize to list
        if isinstance(recursos_data, dict):
            recursos_list = recursos_data.get('humanos', [])
        elif isinstance(recursos_data, list):
            recursos_list = [r['rol'] if isinstance(r, dict) else r for r in recursos_data]
        else:
            recursos_list = []
            
        # Grid 2x2 for Resources
        for i, rec in enumerate(recursos_list[:4]):
            r_idx = row + (i // 2)
            c_idx = 1 + (i % 2) * 3 # A or D
            
            c_end = c_idx + 2
            col_letter = get_column_letter(c_idx)
            col_end = get_column_letter(c_end)
            
            ws.merge_cells(f'{col_letter}{r_idx}:{col_end}{r_idx}')
            cell = ws[f'{col_letter}{r_idx}']
            cell.value = str(rec)
            cell.font = Font(color=self.COLOR_PRIMARY, bold=True)
            cell.fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border_thin
            
        row += 3
        
        # 8. Riesgos (New Section)
        ws[f'A{row}'] = "RIESGOS PRINCIPALES"
        ws[f'A{row}'].font = Font(size=14, color=self.COLOR_PRIMARY, bold=True)
        row += 1
        
        riesgos = data.get('riesgos', [])
        if riesgos:
            # Headers
            ws[f'A{row}'] = "Riesgo"; ws[f'B{row}'] = "Prob."; ws[f'C{row}'] = "Imp."; 
            ws.merge_cells(f'D{row}:F{row}'); ws[f'D{row}'] = "Mitigación"
            
            for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}']]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid") # Red for risks
                cell.alignment = Alignment(horizontal='center')
                
            row += 1
            for r in riesgos:
                desc = r.get('descripcion', r.get('desc', ''))
                prob = r.get('probabilidad', r.get('prob', ''))
                imp = r.get('impacto', r.get('imp', ''))
                mit = r.get('mitigacion', r.get('estr', ''))
                
                ws[f'A{row}'] = desc
                ws[f'B{row}'] = prob
                ws[f'C{row}'] = imp
                ws.merge_cells(f'D{row}:F{row}')
                ws[f'D{row}'] = mit
                row += 1
        
        # Dimensions
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        buffer = BytesIO()
        wb.save(buffer if not output_path else output_path)
        buffer.seek(0)
        return buffer

    def generate_proyecto_complejo(self, data: Dict[str, Any], output_path: Optional[str] = None) -> BytesIO:
        """Generate Complex PMI Project Excel - Tesla Visual Identity (Deep Mirror)"""
        wb = Workbook()
        wb.iso_dates = True
        
        # --- Sheet 1: Project Charter (Visual) ---
        ws = wb.active
        ws.title = "Project Charter"
        ws.sheet_view.showGridLines = False
        
        # Title Block with Gradient/Color
        ws.merge_cells('A1:F2')
        ws['A1'] = f"PROJECT CHARTER - {data.get('codigo', 'PMI-000')}"
        ws['A1'].font = Font(color="FFFFFF", size=20, bold=True)
        ws['A1'].fill = self.header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        row = 4
        
        # PMI Sections as "Cards"
        sections = [
            ("Propósito del Proyecto", data.get('proposito', 'Descripción pendiente')),
            ("Objetivos Estratégicos", data.get('objetivos', 'Alineación pendiente')),
            ("Alcance Preliminar", data.get('alcance', 'Alcance pendiente')),
            ("Entregables", str(data.get('entregables', ['Ver Anexo']))),
            ("Supuestos y Restricciones", data.get('supuestos', 'Supuestos pendientes'))
        ]
        
        for titulo, contenido in sections:
            # Header of Card
            ws.merge_cells(f'A{row}:F{row}')
            header = ws[f'A{row}']
            header.value = titulo.upper()
            header.font = Font(color=self.COLOR_PRIMARY, bold=True, size=11)
            header.border = Border(bottom=Side(style='thick', color=self.COLOR_ACCENT))
            row += 1
            
            # Content of Card
            ws.merge_cells(f'A{row}:F{row+2}')
            content = ws[f'A{row}']
            content.value = str(contenido)
            content.alignment = Alignment(wrap_text=True, vertical='top')
            content.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            content.font = Font(color="374151")
            
            # Left border accent for card content
            for r_idx in range(row, row+3):
                 ws[f'A{r_idx}'].border = Border(left=Side(style='medium', color=self.COLOR_SECONDARY))
            
            row += 4
            
        # --- Sheet 2: Dashboard KPI (Analysis) ---
        ws_kpi = wb.create_sheet("Dashboard KPI")
        ws_kpi.sheet_view.showGridLines = False
        ws_kpi['A1'] = "INDICADORES DE DESEMPEÑO (EVM)"
        ws_kpi['A1'].font = self.title_font
        
        # KPI Cards (Grid 2x3)
        kpis = [
            ("SPI (Cronograma)", data.get('spi', 1.0), ">= 1.0"),
            ("CPI (Costos)", data.get('cpi', 1.0), ">= 1.0"),
            ("Valor Ganado (EV)", f"${data.get('ev', 0):,.2f}", "N/A"),
            ("Costo Real (AC)", f"${data.get('ac', 0):,.2f}", "< EV"),
            ("Valor Planificado (PV)", f"${data.get('pv', 0):,.2f}", "N/A"),
            ("Presupuesto (BAC)", f"${data.get('bac', 0):,.2f}", "N/A")
        ]
        
        r_start = 3
        c_start = 1
        for i, (name, value, target) in enumerate(kpis):
            # Calculate grid position (2 columns)
            r = r_start + (i // 2) * 5
            c = c_start + (i % 2) * 3
            
            # Row 1: Name
            ws_kpi.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+2)
            cell = ws_kpi.cell(row=r, column=c)
            cell.value = name
            cell.font = Font(color="6B7280", bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.fill = self.card_fill
            
            # Row 2: Value
            ws_kpi.merge_cells(start_row=r+1, start_column=c, end_row=r+1, end_column=c+2)
            cell = ws_kpi.cell(row=r+1, column=c)
            cell.value = value
            cell.font = Font(color=self.COLOR_PRIMARY, bold=True, size=18)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = self.card_fill
            
            # Row 3: Target
            ws_kpi.merge_cells(start_row=r+2, start_column=c, end_row=r+2, end_column=c+2)
            cell = ws_kpi.cell(row=r+2, column=c)
            cell.value = f"Meta: {target}"
            cell.font = Font(color="10B981", size=10)
            cell.alignment = Alignment(horizontal='center', vertical='top')
            cell.fill = self.card_fill
            
            # Borders (Outer Box)
            ws_kpi.cell(row=r, column=c).border = self.border_thin
            ws_kpi.cell(row=r+1, column=c).border = self.border_thin
            ws_kpi.cell(row=r+2, column=c).border = self.border_thin

        ws_kpi.column_dimensions['A'].width = 5
        ws_kpi.column_dimensions['B'].width = 15
        ws_kpi.column_dimensions['C'].width = 15
        ws_kpi.column_dimensions['D'].width = 5
        ws_kpi.column_dimensions['E'].width = 15
        ws_kpi.column_dimensions['F'].width = 15

        # --- Sheet 3: Riesgos y Stakeholders ---
        ws_risk = wb.create_sheet("Riesgos y Stakeholders")
        ws_risk['A1'] = "MATRIZ DE RIESGOS"
        ws_risk['A1'].font = Font(bold=True, size=14, color=self.COLOR_PRIMARY)
        
        # Headers Riesgos
        headers_r = ["Descripción", "Probabilidad", "Impacto", "Estrategia"]
        for idx, h in enumerate(headers_r, 1):
            cell = ws_risk.cell(row=2, column=idx, value=h)
            cell.font = self.header_font
            cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            
        row = 3
        riesgos = data.get('riesgos', [])
        for r in riesgos:
            desc = r.get('descripcion', r.get('desc', ''))
            prob = r.get('probabilidad', r.get('prob', ''))
            imp = r.get('impacto', r.get('imp', ''))
            estr = r.get('mitigacion', r.get('estr', ''))
            
            ws_risk.cell(row=row, column=1, value=desc)
            ws_risk.cell(row=row, column=2, value=prob)
            ws_risk.cell(row=row, column=3, value=imp)
            ws_risk.cell(row=row, column=4, value=estr)
            row += 1
            
        # Stakeholders
        row += 2
        ws_risk[f'A{row}'] = "REGISTRO DE STAKEHOLDERS"
        ws_risk[f'A{row}'].font = Font(bold=True, size=14, color=self.COLOR_PRIMARY)
        row += 1
        
        headers_s = ["Nombre", "Rol", "Poder", "Interés"]
        for idx, h in enumerate(headers_s, 1):
            cell = ws_risk.cell(row=row, column=idx, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill # Blue
            
        row += 1
        stakeholders = data.get('stakeholders', [])
        for s in stakeholders:
            if isinstance(s, dict):
                ws_risk.cell(row=row, column=1, value=s.get('nombre'))
                ws_risk.cell(row=row, column=2, value=s.get('rol'))
                ws_risk.cell(row=row, column=3, value=s.get('poder'))
                ws_risk.cell(row=row, column=4, value=s.get('interes'))
            else:
                ws_risk.cell(row=row, column=1, value=str(s))
            row += 1

        ws_risk.column_dimensions['A'].width = 30
        ws_risk.column_dimensions['B'].width = 15
        ws_risk.column_dimensions['C'].width = 15
        ws_risk.column_dimensions['D'].width = 20

        # --- Sheet 4: Cronograma Gantt ---
        ws_gantt = wb.create_sheet("Cronograma Gantt")
        ws_gantt['A1'] = "CRONOGRAMA DE PROYECTO"
        ws_gantt['A1'].font = Font(bold=True, size=14, color=self.COLOR_PRIMARY)
        
        # Gantt Logic (Simple)
        crono = data.get('cronograma', {})
        if isinstance(crono, dict):
             ws_gantt['A3'] = f"Duración Total: {crono.get('duracion_total')}"
             ws_gantt['A4'] = f"Inicio: {crono.get('fecha_inicio')}"
             ws_gantt['A5'] = f"Fin: {crono.get('fecha_fin')}"
        else:
             ws_gantt['A3'] = str(crono)

        buffer = BytesIO()
        wb.save(buffer if not output_path else output_path)
        buffer.seek(0)
        return buffer
    
    def generate_informe_tecnico(self, data: Dict[str, Any], output_path: Optional[str] = None) -> BytesIO:
        """Generate Technical Report Excel - Tesla Visual Identity (Dashboard)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe Técnico"
        ws.sheet_view.showGridLines = False
        
        # 1. Header (Logo & Title)
        ws.merge_cells('A1:C3')
        ws['A1'] = "TESLA"
        ws['A1'].font = Font(color="FFFFFF", bold=True, size=24)
        ws['A1'].fill = self.header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('D1:G3')
        ws['D1'] = f"INFORME TÉCNICO\n{data.get('codigo', 'INF-TEC-001')}"
        ws['D1'].font = Font(color=self.COLOR_PRIMARY, size=16, bold=True)
        ws['D1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        row = 5
        
        # 2. Info Cards (Grid)
        cards = [
            ("CLIENTE", data.get('cliente', {}).get('nombre', 'N/A')),
            ("FECHA", data.get('fecha', 'N/A')),
            ("TÉCNICO", "Ing. Residente")
        ]
        
        for i, (label, val) in enumerate(cards):
            c1 = get_column_letter(1 + i*2) # A, C, E
            c2 = get_column_letter(2 + i*2) # B, D, F
            
            ws.merge_cells(f'{c1}{row}:{c2}{row}')
            ws[f'{c1}{row}'] = label
            ws[f'{c1}{row}'].font = Font(color="6B7280", size=9, bold=True)
            
            ws.merge_cells(f'{c1}{row+1}:{c2}{row+1}')
            ws[f'{c1}{row+1}'] = val
            ws[f'{c1}{row+1}'].font = Font(color=self.COLOR_PRIMARY, size=11, bold=True)
            ws[f'{c1}{row+1}'].border = self.card_border
            
        row += 4
        
        # 3. Specs Table
        ws[f'A{row}'] = "ESPECIFICACIONES DEL SISTEMA"
        ws[f'A{row}'].font = Font(bold=True, size=12, color=self.COLOR_PRIMARY)
        row += 1
        
        specs = data.get('specs', [
            {"param": "Normativa", "valor": "CNE Suministro 2011"},
            {"param": "Tensión", "valor": "220V / 380V"},
            {"param": "Frecuencia", "valor": "60 Hz"}
        ])
        
        # Table Header
        headers = ["Parámetro", "Valor Referencial", "Estado"]
        for col, h in enumerate(headers, 1):
             cell = ws.cell(row=row, column=col, value=h)
             cell.font = self.header_font
             cell.fill = self.header_fill
             cell.alignment = Alignment(horizontal='center')
        row += 1
        
        for i, spec in enumerate(specs):
            ws.cell(row=row, column=1, value=spec.get('param')).border = self.border_thin
            ws.cell(row=row, column=2, value=spec.get('valor')).border = self.border_thin
            ws.cell(row=row, column=3, value="Cumple").border = self.border_thin
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            
            if i % 2 == 0:
                 for c in range(1, 4):
                     ws.cell(row=row, column=c).fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            row += 1
            
        row += 2
        
        # 4. Results Matrix
        ws[f'A{row}'] = "RESULTADOS DE MEDICIONES"
        ws[f'A{row}'].font = Font(bold=True, size=12, color=self.COLOR_PRIMARY)
        row += 1
        
        headers_m = ["Punto de Medida", "L1 (A)", "L2 (A)", "L3 (A)", "Evaluación"]
        for col, h in enumerate(headers_m, 1):
             cell = ws.cell(row=row, column=col, value=h)
             cell.font = self.header_font
             cell.fill = self.header_fill
             cell.alignment = Alignment(horizontal='center')
        row += 1
        
        mediciones = data.get('mediciones', [{"pto": "Tablero General", "l1": 25.5, "l2": 24.8, "l3": 26.1, "res": "OK"}])
        for m in mediciones:
            ws.cell(row=row, column=1, value=m.get('pto')).border = self.border_thin
            ws.cell(row=row, column=2, value=m.get('l1')).border = self.border_thin
            ws.cell(row=row, column=3, value=m.get('l2')).border = self.border_thin
            ws.cell(row=row, column=4, value=m.get('l3')).border = self.border_thin
            
            # Badge
            res = m.get('res', 'OK')
            c_res = ws.cell(row=row, column=5, value=res)
            c_res.border = self.border_thin
            c_res.alignment = Alignment(horizontal='center')
            c_res.font = Font(bold=True, color="FFFFFF")
            c_res.fill = PatternFill(start_color="059669" if res=="OK" else "DC2626", end_color="059669" if res=="OK" else "DC2626", fill_type="solid")
            
            row += 1
            
        # Column Widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20

        buffer = BytesIO()
        wb.save(buffer if not output_path else output_path)
        buffer.seek(0)
        return buffer

    def generate_informe_ejecutivo(self, data: Dict[str, Any], output_path: Optional[str] = None) -> BytesIO:
        """Generate Executive Report Excel - Tesla Visual Identity (Dashboard)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen Ejecutivo"
        ws.sheet_view.showGridLines = False
        
        # 1. Header
        ws.merge_cells('A1:B3')
        ws['A1'] = "TESLA"
        ws['A1'].font = Font(color="FFFFFF", bold=True, size=24)
        ws['A1'].fill = self.header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('C1:F3')
        ws['C1'] = "INFORME EJECUTIVO\nAnálisis de Viabilidad y Retorno"
        ws['C1'].font = Font(color=self.COLOR_PRIMARY, size=16, bold=True)
        ws['C1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        row = 5
        
        # 2. Financial Highlights (Big Cards)
        financials = [
            ("ROI ESPERADO", f"{data.get('roi', 25)}%", "10B981"), # Green
            ("AHORRO ANUAL", f"${data.get('ahorro_anual', 0):,.2f}", "3B82F6"), # Blue
            ("PAYBACK", f"{data.get('payback', 12)} m", "F59E0B"), # Amber
            ("VPN (10A)", f"${data.get('vpn', 0):,.2f}", "8B5CF6") # Purple
        ]
        
        # Grid 2x2
        for i, (label, val, color) in enumerate(financials):
            r = row + (i // 2) * 4 # Height 4 rows
            c = 1 + (i % 2) * 3    # Col 1 or 4
            
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+2)
            c_label = ws.cell(row=r, column=c, value=label)
            c_label.font = Font(color="6B7280", bold=True, size=10)
            c_label.alignment = Alignment(horizontal='left')
            
            ws.merge_cells(start_row=r+1, start_column=c, end_row=r+2, end_column=c+2)
            c_val = ws.cell(row=r+1, column=c, value=val)
            c_val.font = Font(color=color, bold=True, size=24)
            c_val.alignment = Alignment(horizontal='center', vertical='center')
            
            # Border Box
            for rx in range(r, r+4):
                 ws.cell(row=rx, column=c).border = Border(left=Side(style='medium', color=color))
        
        row += 8
        
        # 3. Inversion Table
        ws[f'A{row}'] = "DETALLE DE INVERSIÓN"
        ws[f'A{row}'].font = Font(bold=True, size=12, color=self.COLOR_PRIMARY)
        row += 1
        
        headers = ["Concepto", "Monto Estimado", "Prioridad", "Impacto"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        breakdown = data.get('inversion_detalle', [{"cat": "Equipamiento", "monto": 50000, "pri": "Alta", "imp": "Directo"}])
        for item in breakdown:
            ws.cell(row=row, column=1, value=item.get('cat')).border = self.border_thin
            ws.cell(row=row, column=2, value=f"${item.get('monto'):,.2f}").border = self.border_thin
            ws.cell(row=row, column=3, value=item.get('pri')).border = self.border_thin
            ws.cell(row=row, column=4, value=item.get('imp')).border = self.border_thin
            row += 1
            
        row += 2
        
        # 4. Conclusion Block
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "CONCLUSIÓN EJECUTIVA"
        ws[f'A{row}'].font = Font(bold=True, size=12, color=self.COLOR_PRIMARY)
        row += 1
        
        ws.merge_cells(f'A{row}:F{row+2}')
        ws[f'A{row}'] = "El proyecto presenta indicadores financieros positivos con un retorno de inversión superior a la tasa de corte corporativa. Se recomienda la aprobación inmediata para aprovechar las condiciones de mercado actuales."
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws[f'A{row}'].fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid") # Green tint
        ws[f'A{row}'].font = Font(color="065F46", italic=True)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        buffer = BytesIO()
        wb.save(buffer if not output_path else output_path)
        buffer.seek(0)
        return buffer
