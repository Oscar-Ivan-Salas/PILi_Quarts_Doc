import sys
import os
import glob
from pathlib import Path
from openpyxl import load_workbook

# Colors for terminal
GREEN = '\033[92m'
RED = '\033[91m'
RESET = '\033[0m'

def check_18_of_truth():
    print(f"\n{GREEN}[PROTOCOL] INSPECCIÓN DE LOS 18 DE LA VERDAD{RESET}")
    print("==================================================")
    
    storage_dir = Path(r"e:\PILi_Quarts\workspace-modern\storage\generados")
    
    # 1. Identify the latest batch timestamp
    # We saw files with 1771078754 and 1771078755. Let's look for *177107875*
    pattern = "177107875"
    # Use os.listdir for safety
    all_files = os.listdir(storage_dir)
    files = [storage_dir / f for f in all_files if pattern in f]
    
    if len(files) < 18:
        print(f"{RED}[ALERTA] Se encontraron solo {len(files)} archivos con el patrón {pattern}. Se esperaban 18.{RESET}")
        # List what we found
        for f in files: print(f" - {f.name}")
        return 1

    print(f"Archivos Identificados del Lote {pattern}: {len(files)}")
    
    # Organize by Type
    cot_compleja_xlsx = None
    
    files_by_type = {}
    
    for f in files:
        ext = f.suffix.lower()
        name = f.stem
        # Simplified type detection
        if "cotizacion_compleja" in name: 
            type_key = "Cotización Compleja"
            if ext == ".xlsx": cot_compleja_xlsx = f
        elif "cotizacion_simple" in name: type_key = "Cotización Simple"
        elif "proyecto_complejo" in name: type_key = "Proyecto Complejo"
        elif "proyecto_simple" in name: type_key = "Proyecto Simple"
        elif "informe_tecnico" in name: type_key = "Informe Técnico"
        elif "informe_ejecutivo" in name: type_key = "Informe Ejecutivo"
        else: type_key = "Otro"
        
        if type_key not in files_by_type: files_by_type[type_key] = []
        files_by_type[type_key].append(f)

    # 2. List and Verify Existence
    all_exist = True
    print("\n[INVENTARIO]")
    for t, fs in files_by_type.items():
        print(f"  > {t}: {len(fs)} archivos")
        for f in fs:
            print(f"    - {f.name} ({f.stat().st_size} bytes)")
    
    # 3. Deep Verify Excel (Cotización Compleja)
    print("\n[AUDITORÍA PROFUNDA: EXCEL]")
    if cot_compleja_xlsx:
        try:
            # Check Formula
            wb_formula = load_workbook(cot_compleja_xlsx, data_only=False)
            ws_f = wb_formula.active
            # Assuming Totals are at the bottom. 
            # In excel_generator, we write totals. 
            # Subtotal, IGV, Total are usually in the last rows of the totals section.
            # Let's search for "TOTAL:" cell and check the value next to it.
            
            total_label_cell = None
            for row in ws_f.iter_rows():
                for cell in row:
                    if cell.value == "TOTAL:":
                        total_label_cell = cell
                        break
                if total_label_cell: break
            
            if total_label_cell:
                # Value is likely in the next column or +1
                # In generator: table with data [['TOTAL:', total_val]] -> col 1 is label, col 2 is value
                # But table writes to cells.
                # Let's assume it's the cell to the right.
                val_cell = ws_f.cell(row=total_label_cell.row, column=total_label_cell.column + 1)
                
                print(f"  > Celda TOTAL encontrada en {total_label_cell.coordinate}")
                print(f"  > Contenido (Fórmula/Valor Rapido): {val_cell.value}")
                
                # Check calculation (Data Only)
                # wb_data = load_workbook(cot_compleja_xlsx, data_only=True)
                # ws_d = wb_data.active
                # val_cell_d = ws_d.cell(row=total_label_cell.row, column=total_label_cell.column + 1)
                # print(f"  > Valor Calculado: {val_cell_d.value}")
                
                # Verify Logic
                # expected = 1156.40
                # if abs(float(val_cell_d.value or 0) - expected) < 0.01:
                #    print(f"  {GREEN}✅ CONSISTENCIA CONFIRMADA: S/ {val_cell_d.value}{RESET}")
                # else:
                #    print(f"  {RED}❌ DISCREPANCIA: Esperado {expected}, Encontrado {val_cell_d.value}{RESET}")
                
                if "=" in str(val_cell.value):
                     print(f"  {GREEN}✅ FÓRMULA DETECTADA (Excel Vivo){RESET}")
                else:
                     print(f"  {RED}⚠️ ALERTA: No se detectó fórmula explícita (Valor estático?){RESET}")

            else:
                print(f"{RED}❌ No se encontró la etiqueta 'TOTAL:' en el Excel.{RESET}")
        except Exception as e:
            print(f"{RED}❌ Error leyendo Excel: {e}{RESET}")
    else:
        print(f"{RED}❌ No se encontró el Excel de Cotización Compleja para auditar.{RESET}")

    # 4. Final Verdict
    print("\n--------------------------------------------------")
    print("Enlaces de Descarga (Rutas Absolutas):")
    for f in files:
        print(f"file:///{f.absolute().as_posix()}")
    print("--------------------------------------------------")
    
    return 0

if __name__ == "__main__":
    check_18_of_truth()
